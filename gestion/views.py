import re
from django.contrib import messages
from django.db import connection
from django.shortcuts import redirect, render
from accounts.decorators import superadmin_required, admin_year_required
import os 
from django.conf import settings
from django.db import connection, IntegrityError, DatabaseError
import openpyxl
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.utils import timezone
from django.utils.safestring import mark_safe
import openpyxl
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.http import HttpResponse
from django.db import connection, IntegrityError, DatabaseError
from django.utils.safestring import mark_safe
import openpyxl
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.http import HttpResponse
from django.db import connection, IntegrityError, DatabaseError

# ---------- HELPERS ----------
def fetchall_dict(cur):
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


# ---------- DASHBOARD ----------# views.py
import json
from django.db import connection
from django.shortcuts import render
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime, timedelta

@superadmin_required
def dashboard(request):
    with connection.cursor() as cur:
        # Statistiques principales
        cur.execute("SELECT COUNT(*) FROM isms.etudiant;")
        nb_etudiants = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM isms.responsable;")
        nb_responsables = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM isms.memoire;")
        total_memoires = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM isms.soutenance;")
        total_soutenances = cur.fetchone()[0]
        
        # Mémoires par statut
        cur.execute("""
            SELECT statut, COUNT(*) 
            FROM isms.memoire 
            GROUP BY statut 
            ORDER BY statut;
        """)
        mem_stats = cur.fetchall()
        
        # Soutenances par mois (derniers 6 mois) - CORRECTION ICI
        cur.execute("""
            SELECT 
                TO_CHAR(date_, 'Mon YYYY') as mois,
                COUNT(*) as total
            FROM isms.soutenance
            WHERE date_ >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY TO_CHAR(date_, 'Mon YYYY'), 
                     DATE_TRUNC('month', date_)
            ORDER BY DATE_TRUNC('month', date_);
        """)
        soutenances_par_mois = cur.fetchall()
        
        # Étudiants par département
        cur.execute("""
            SELECT d.nom_departement, COUNT(e.id_etudiant) as total
            FROM isms.departement d
            LEFT JOIN isms.etudiant e ON d.id_departement = e.id_departement
            GROUP BY d.nom_departement
            ORDER BY total DESC;
        """)
        etudiants_par_dept = cur.fetchall()
        
        # Statistiques additionnelles
        cur.execute("SELECT COUNT(*) FROM isms.jury;")
        nb_jurys = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(*) FROM isms.salle;")
        nb_salles = cur.fetchone()[0]
        
        cur.execute("SELECT COUNT(DISTINCT id_departement) FROM isms.departement;")
        nb_departements = cur.fetchone()[0]
        
        cur.execute("""
            SELECT COUNT(*) 
            FROM isms.memoire 
            WHERE statut = 'VALIDE';
        """)
        memoires_valides = cur.fetchone()[0]
        
        cur.execute("""
            SELECT COUNT(*) 
            FROM isms.memoire 
            WHERE statut = 'EN_VERIFICATION';
        """)
        memoires_en_verification = cur.fetchone()[0]
        
        # Mémoires par type
        cur.execute("""
            SELECT type, COUNT(*) as total
            FROM isms.memoire
            GROUP BY type
            ORDER BY total DESC;
        """)
        memoires_par_type = cur.fetchall()
        
        # Distribution des notes - VERSION CORRIGÉE ET SIMPLIFIÉE
        cur.execute("""
            SELECT tranche, COUNT(*) as total
            FROM (
                SELECT 
                    CASE 
                        WHEN note_finale < 10 THEN 'Moins de 10'
                        WHEN note_finale >= 10 AND note_finale < 12 THEN '10 - 12'
                        WHEN note_finale >= 12 AND note_finale < 14 THEN '12 - 14'
                        WHEN note_finale >= 14 AND note_finale < 16 THEN '14 - 16'
                        ELSE 'Plus de 16'
                    END as tranche,
                    CASE 
                        WHEN note_finale < 10 THEN 1
                        WHEN note_finale >= 10 AND note_finale < 12 THEN 2
                        WHEN note_finale >= 12 AND note_finale < 14 THEN 3
                        WHEN note_finale >= 14 AND note_finale < 16 THEN 4
                        ELSE 5
                    END as ordre
                FROM isms.note
                WHERE note_finale IS NOT NULL
            ) AS notes_classees
            GROUP BY tranche, ordre
            ORDER BY ordre;
        """)
        notes_distribution = cur.fetchall()
        
        # Activité des 7 derniers jours
        cur.execute("""
            SELECT 
                TO_CHAR(date_depot, 'DD/MM') as jour,
                COUNT(*) as total
            FROM isms.memoire
            WHERE date_depot >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY date_depot, TO_CHAR(date_depot, 'DD/MM')
            ORDER BY date_depot;
        """)
        activite_7_jours = cur.fetchall()
        
        # Si pas de données pour les 7 derniers jours, créer des données vides
        if not activite_7_jours:
            activite_7_jours = []
            for i in range(7):
                jour = datetime.now() - timedelta(days=6-i)
                activite_7_jours.append((jour.strftime('%d/%m'), 0))
    
    # Fonction helper pour convertir les tuples en format JSON
    def format_data(data):
        return [{'label': str(row[0]), 'value': int(row[1])} for row in data]
    
    # Convertir les données en JSON pour JavaScript
    context = {
        # Statistiques principales
        'nb_etudiants': nb_etudiants,
        'nb_responsables': nb_responsables,
        'total_memoires': total_memoires,
        'total_soutenances': total_soutenances,
        
        # Statistiques additionnelles
        'nb_jurys': nb_jurys,
        'nb_salles': nb_salles,
        'nb_departements': nb_departements,
        'memoires_valides': memoires_valides,
        'memoires_en_verification': memoires_en_verification,
        
        # Données pour les graphiques (en JSON)
        'mem_stats_json': json.dumps(format_data(mem_stats), cls=DjangoJSONEncoder),
        'soutenances_par_mois_json': json.dumps(format_data(soutenances_par_mois), cls=DjangoJSONEncoder),
        'etudiants_par_dept_json': json.dumps(format_data(etudiants_par_dept), cls=DjangoJSONEncoder),
        'memoires_par_type_json': json.dumps(format_data(memoires_par_type), cls=DjangoJSONEncoder),
        'notes_distribution_json': json.dumps(format_data(notes_distribution), cls=DjangoJSONEncoder),
        'activite_7_jours_json': json.dumps(format_data(activite_7_jours), cls=DjangoJSONEncoder),
    }
    
    return render(request, "gestion/dashboard.html", context)

# =========================
# REFERENTIELS
# =========================

# ---- ANNEE UNIVERSITAIRE ----

@superadmin_required
def annee_select(request):
    """Liste des années + choix (stocké en session)."""
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id_annee, libelle, active
            FROM isms.annee_universitaire
            ORDER BY id_annee DESC;
        """)
        rows = fetchall_dict(cur)

    return render(request, "gestion/annee_select.html", {"rows": rows})


@superadmin_required
def annee_set_session(request, pk):
    """Sélectionner une année : on la met en session (et optionnellement active en DB)."""
    with connection.cursor() as cur:
        cur.execute("SELECT libelle FROM isms.annee_universitaire WHERE id_annee=%s;", [pk])
        row = cur.fetchone()

    if not row:
        messages.error(request, "Année introuvable.")
        return redirect("gestion:annee_select")

    libelle = row[0]
    request.session["annee_id"] = int(pk)
    request.session["annee_libelle"] = libelle

    # OPTION (recommandée) : rendre l'année active globalement
    with connection.cursor() as cur:
        cur.execute("UPDATE isms.annee_universitaire SET active = FALSE;")
        cur.execute("UPDATE isms.annee_universitaire SET active = TRUE WHERE id_annee=%s;", [pk])

    messages.success(request, f"Année sélectionnée : {libelle}")
    return redirect("gestion:dashboard")


@superadmin_required
def annee_create_and_select(request):
    """Créer une année + la sélectionner immédiatement."""
    if request.method != "POST":
        return redirect("gestion:annee_select")

    libelle = (request.POST.get("libelle") or "").strip()
    if not libelle:
        messages.error(request, "Libellé obligatoire.")
        return redirect("gestion:annee_select")

    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO isms.annee_universitaire(libelle, active)
            VALUES (%s, TRUE)
            RETURNING id_annee;
        """, [libelle])
        new_id = cur.fetchone()[0]

        # garantir une seule active
        cur.execute("UPDATE isms.annee_universitaire SET active = FALSE WHERE id_annee <> %s;", [new_id])

    request.session["annee_id"] = int(new_id)
    request.session["annee_libelle"] = libelle

    messages.success(request, f"Année créée et sélectionnée : {libelle}")
    return redirect("gestion:dashboard")


@superadmin_required
def annee_list(request):
    with connection.cursor() as cur:
        cur.execute("SELECT id_annee, libelle, active FROM isms.annee_universitaire ORDER BY id_annee DESC;")
        rows = fetchall_dict(cur)
    return render(request, "gestion/annees/list.html", {"rows": rows})


@superadmin_required
def annee_create(request):
    if request.method == "POST":
        libelle = (request.POST.get("libelle") or "").strip()
        if not libelle:
            messages.error(request, "Libellé obligatoire.")
            return redirect("gestion:annee_list")

        with connection.cursor() as cur:
            cur.execute("INSERT INTO isms.annee_universitaire(libelle, active) VALUES (%s, FALSE);", [libelle])

        messages.success(request, "Année ajoutée.")
        return redirect("gestion:annee_list")

    return render(request, "gestion/annees/create.html")


@superadmin_required
def annee_activate(request, pk):
    # une seule active (index unique partiel)
    with connection.cursor() as cur:
        cur.execute("UPDATE isms.annee_universitaire SET active = FALSE;")
        cur.execute("UPDATE isms.annee_universitaire SET active = TRUE WHERE id_annee = %s;", [pk])
    messages.success(request, "Année activée.")
    return redirect("gestion:annee_list")


@superadmin_required
def annee_delete(request, pk):
    with connection.cursor() as cur:
        cur.execute("DELETE FROM isms.annee_universitaire WHERE id_annee = %s;", [pk])
    messages.success(request, "Année supprimée.")
    return redirect("gestion:annee_list")


# ---- NIVEAU ----
@superadmin_required
def niveau_list(request):
    with connection.cursor() as cur:
        cur.execute("SELECT id_niveau, libelle FROM isms.niveau ORDER BY libelle;")
        rows = fetchall_dict(cur)
    return render(request, "gestion/niveaux/list.html", {"rows": rows})


@superadmin_required
def niveau_create(request):
    if request.method == "POST":
        libelle = (request.POST.get("libelle") or "").strip()
        if not libelle:
            messages.error(request, "Libellé obligatoire.")
            return redirect("gestion:niveau_list")
        with connection.cursor() as cur:
            cur.execute("INSERT INTO isms.niveau(libelle) VALUES (%s);", [libelle])
        messages.success(request, "Niveau ajouté.")
        return redirect("gestion:niveau_list")

    return render(request, "gestion/niveaux/create.html")

@superadmin_required
def niveau_update(request, pk):
    with connection.cursor() as cur:
        cur.execute("SELECT id_niveau, libelle FROM isms.niveau WHERE id_niveau=%s;", [pk])
        row = cur.fetchone()

    if not row:
        messages.error(request, "Niveau introuvable.")
        return redirect("gestion:niveau_list")

    niveau = {"id_niveau": row[0], "libelle": row[1]}

    if request.method == "POST":
        libelle = (request.POST.get("libelle") or "").strip()
        if not libelle:
            messages.error(request, "Libellé obligatoire.")
            return redirect("gestion:niveau_update", pk=pk)

        with connection.cursor() as cur:
            cur.execute("UPDATE isms.niveau SET libelle=%s WHERE id_niveau=%s;", [libelle, pk])

        messages.success(request, "Niveau modifié.")
        return redirect("gestion:niveau_list")

    return render(request, "gestion/niveaux/update.html", {"niveau": niveau})

from django.db import IntegrityError

@superadmin_required
def niveau_delete(request, pk):
    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.niveau WHERE id_niveau = %s;", [pk])
        messages.success(request, "Niveau supprimé.")
    except IntegrityError:
        messages.error(request, "Impossible de supprimer : ce niveau est lié à d'autres données (ex: départements).")
    return redirect("gestion:niveau_list")


# ---- DEPARTEMENT ----@superadmin_required
def departement_list(request):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT d.id_departement, d.nom_departement, d.id_niveau, n.libelle AS niveau
            FROM isms.departement d
            JOIN isms.niveau n ON n.id_niveau = d.id_niveau
            ORDER BY d.nom_departement;
        """)
        rows = fetchall_dict(cur)

        cur.execute("SELECT id_niveau, libelle FROM isms.niveau ORDER BY libelle;")
        niveaux = fetchall_dict(cur)

    return render(request, "gestion/departements/list.html", {
        "rows": rows,
        "niveaux": niveaux
    })


@superadmin_required
def departement_create(request):
    if request.method == "POST":
        nom = (request.POST.get("nom_departement") or "").strip()
        id_niveau = request.POST.get("id_niveau")
        if not nom or not id_niveau:
            messages.error(request, "Nom + niveau obligatoires.")
            return redirect("gestion:departement_list")

        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO isms.departement(nom_departement, id_niveau) VALUES (%s, %s);",
                [nom, id_niveau],
            )
        messages.success(request, "Département ajouté.")
        return redirect("gestion:departement_list")

    return redirect("gestion:departement_list")


from django.db import IntegrityError

@superadmin_required
def departement_delete(request, pk):
    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.departement WHERE id_departement = %s;", [pk])
        messages.success(request, "Département supprimé.")
    except IntegrityError:
        # lié à d'autres tables (ex: etudiant, memoire, etc.)
        messages.error(request, "Impossible de supprimer : ce département est lié à d'autres données.")
    return redirect("gestion:departement_list")

@superadmin_required
def departement_update(request, pk):
    with connection.cursor() as cur:
        # charger departement
        cur.execute("""
            SELECT id_departement, nom_departement, id_niveau
            FROM isms.departement
            WHERE id_departement = %s;
        """, [pk])
        dep = cur.fetchone()

        # charger niveaux
        cur.execute("SELECT id_niveau, libelle FROM isms.niveau ORDER BY libelle;")
        niveaux = fetchall_dict(cur)

    if not dep:
        messages.error(request, "Département introuvable.")
        return redirect("gestion:departement_list")

    dep_data = {
        "id_departement": dep[0],
        "nom_departement": dep[1],
        "id_niveau": dep[2],
    }

    if request.method == "POST":
        nom = (request.POST.get("nom_departement") or "").strip()
        id_niveau = request.POST.get("id_niveau")

        if not nom or not id_niveau:
            messages.error(request, "Nom + niveau obligatoires.")
            return redirect("gestion:departement_update", pk=pk)

        with connection.cursor() as cur:
            cur.execute("""
                UPDATE isms.departement
                SET nom_departement=%s, id_niveau=%s
                WHERE id_departement=%s;
            """, [nom, id_niveau, pk])

        messages.success(request, "Département modifié.")
        return redirect("gestion:departement_list")

    return render(request, "gestion/departements/update.html", {
        "dep": dep_data,
        "niveaux": niveaux
    })


# ---- SALLE ----
@superadmin_required
def salle_list(request):
    with connection.cursor() as cur:
        cur.execute("SELECT id_salle, nom_salle FROM isms.salle ORDER BY nom_salle;")
        rows = fetchall_dict(cur)
    return render(request, "gestion/salles/list.html", {"rows": rows})



@superadmin_required
def salle_create(request):
    if request.method == "POST":
        nom = (request.POST.get("nom_salle") or "").strip()
        if not nom:
            messages.error(request, "Nom obligatoire.")
            return redirect("gestion:salle_list")
        with connection.cursor() as cur:
            cur.execute("INSERT INTO isms.salle(nom_salle) VALUES (%s);", [nom])
        messages.success(request, "Salle ajoutée.")
        return redirect("gestion:salle_list")
    return render(request, "gestion/salles/create.html")

@superadmin_required
def salle_update(request, pk):
    with connection.cursor() as cur:
        cur.execute("SELECT id_salle, nom_salle FROM isms.salle WHERE id_salle=%s;", [pk])
        row = cur.fetchone()

    if not row:
        messages.error(request, "Salle introuvable.")
        return redirect("gestion:salle_list")

    salle = {"id_salle": row[0], "nom_salle": row[1]}

    if request.method == "POST":
        nom = (request.POST.get("nom_salle") or "").strip()
        if not nom:
            messages.error(request, "Nom obligatoire.")
            return redirect("gestion:salle_update", pk=pk)

        with connection.cursor() as cur:
            cur.execute("UPDATE isms.salle SET nom_salle=%s WHERE id_salle=%s;", [nom, pk])

        messages.success(request, "Salle modifiée.")
        return redirect("gestion:salle_list")

    return render(request, "gestion/salles/update.html", {"salle": salle})


from django.db import IntegrityError

@superadmin_required
def salle_delete(request, pk):
    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.salle WHERE id_salle = %s;", [pk])
        messages.success(request, "Salle supprimée.")
    except IntegrityError:
        messages.error(request, "Impossible de supprimer : cette salle est liée à d'autres données (ex: soutenances).")
    return redirect("gestion:salle_list")


# ---- ROLE ----
@superadmin_required
def role_list(request):
    with connection.cursor() as cur:
        cur.execute("SELECT id_role, code, libelle FROM isms.role ORDER BY code;")
        rows = fetchall_dict(cur)
    return render(request, "gestion/roles/list.html", {"rows": rows})


@superadmin_required
def role_create(request):
    if request.method == "POST":
        code = (request.POST.get("code") or "").strip().upper()
        libelle = (request.POST.get("libelle") or "").strip()
        if not code or not libelle:
            messages.error(request, "Code + libellé obligatoires.")
            return redirect("gestion:role_list")
        with connection.cursor() as cur:
            cur.execute("INSERT INTO isms.role(code, libelle) VALUES (%s, %s);", [code, libelle])
        messages.success(request, "Rôle ajouté.")
        return redirect("gestion:role_list")
    return render(request, "gestion/roles/create.html")

@superadmin_required
def role_update(request, pk):
    with connection.cursor() as cur:
        cur.execute("SELECT id_role, code, libelle FROM isms.role WHERE id_role=%s;", [pk])
        row = cur.fetchone()

    if not row:
        messages.error(request, "Rôle introuvable.")
        return redirect("gestion:role_list")

    role = {"id_role": row[0], "code": row[1], "libelle": row[2]}

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip().upper()
        libelle = (request.POST.get("libelle") or "").strip()

        if not code or not libelle:
            messages.error(request, "Code + libellé obligatoires.")
            return redirect("gestion:role_update", pk=pk)

        with connection.cursor() as cur:
            cur.execute("""
                UPDATE isms.role
                SET code=%s, libelle=%s
                WHERE id_role=%s;
            """, [code, libelle, pk])

        messages.success(request, "Rôle modifié.")
        return redirect("gestion:role_list")

    return render(request, "gestion/roles/update.html", {"role": role})

@superadmin_required
def role_delete(request, pk):
    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.role WHERE id_role = %s;", [pk])
        messages.success(request, "Rôle supprimé.")
    except IntegrityError:
        messages.error(request, "Impossible de supprimer : ce rôle est utilisé (ex: responsables).")
    return redirect("gestion:role_list")


@superadmin_required
def role_delete(request, pk):
    with connection.cursor() as cur:
        cur.execute("DELETE FROM isms.role WHERE id_role = %s;", [pk])
    messages.success(request, "Rôle supprimé.")
    return redirect("gestion:role_list")


# =========================
# UTILISATEURS: RESPONSABLE
# =========================

# =========================
# UTILISATEURS: RESPONSABLE (avec rôles multiples)
# =========================

@superadmin_required
def responsable_list(request):
    with connection.cursor() as cur:
        # Récupérer responsables avec leurs rôles (agrégés)
        cur.execute("""
            SELECT 
                r.id_responsable, 
                r.nom, 
                r.prenom, 
                r.email, 
                r.is_admin,
                COALESCE(
                    array_agg(ro.code ORDER BY ro.code) FILTER (WHERE ro.code IS NOT NULL),
                    '{}'
                ) AS role_codes
            FROM isms.responsable r
            LEFT JOIN isms.responsable_role rr ON rr.id_responsable = r.id_responsable
            LEFT JOIN isms.role ro ON ro.id_role = rr.id_role
            GROUP BY r.id_responsable, r.nom, r.prenom, r.email, r.is_admin
            ORDER BY r.id_responsable DESC;
        """)
        rows = fetchall_dict(cur)

        # Tous les rôles disponibles
        cur.execute("SELECT id_role, code, libelle FROM isms.role ORDER BY code;")
        roles = fetchall_dict(cur)

    return render(request, "gestion/responsables/list.html", {
        "rows": rows, 
        "roles": roles
    })


@superadmin_required
def responsable_create(request):
    if request.method == "POST":
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        mot_de_pass = (request.POST.get("mot_de_pass") or "").strip()
        is_admin = True if request.POST.get("is_admin") == "on" else False
        
        # ✅ NOUVEAU : récupérer les rôles sélectionnés (checkbox multiple)
        role_ids = request.POST.getlist("role_ids")

        if not (nom and prenom and email and mot_de_pass):
            messages.error(request, "Nom, prénom, email et mot de passe sont obligatoires.")
            return redirect("gestion:responsable_list")

        if not role_ids:
            messages.error(request, "Veuillez sélectionner au moins un rôle.")
            return redirect("gestion:responsable_list")

        try:
            with connection.cursor() as cur:
                # Insérer le responsable
                cur.execute("""
                    INSERT INTO isms.responsable(nom, prenom, email, mot_de_pass, is_admin)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id_responsable;
                """, [nom, prenom, email, mot_de_pass, is_admin])
                
                new_id = cur.fetchone()[0]
                
                # ✅ Insérer les rôles dans responsable_role
                for role_id in role_ids:
                    cur.execute("""
                        INSERT INTO isms.responsable_role(id_responsable, id_role)
                        VALUES (%s, %s);
                    """, [new_id, role_id])

            messages.success(request, "Responsable créé avec ses rôles.")
        except IntegrityError as e:
            if 'email' in str(e).lower():
                messages.error(request, "Cet email est déjà utilisé.")
            else:
                messages.error(request, "Erreur : données invalides.")
        except DatabaseError as e:
            messages.error(request, f"Erreur base de données : {str(e)}")

        return redirect("gestion:responsable_list")

    return redirect("gestion:responsable_list")


@superadmin_required
def responsable_update(request, pk):
    with connection.cursor() as cur:
        # Responsable
        cur.execute("""
            SELECT id_responsable, nom, prenom, email, is_admin
            FROM isms.responsable
            WHERE id_responsable=%s;
        """, [pk])
        row = cur.fetchone()

        # Tous les rôles
        cur.execute("SELECT id_role, code, libelle FROM isms.role ORDER BY code;")
        roles = fetchall_dict(cur)
        
        # Rôles actuels du responsable
        cur.execute("""
            SELECT id_role
            FROM isms.responsable_role
            WHERE id_responsable=%s;
        """, [pk])
        current_roles = [r[0] for r in cur.fetchall()]

    if not row:
        messages.error(request, "Responsable introuvable.")
        return redirect("gestion:responsable_list")

    resp = {
        "id_responsable": row[0],
        "nom": row[1],
        "prenom": row[2],
        "email": row[3],
        "is_admin": row[4],
    }

    if request.method == "POST":
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        is_admin = True if request.POST.get("is_admin") == "on" else False
        new_pass = (request.POST.get("mot_de_pass") or "").strip()
        
        # ✅ Récupérer les nouveaux rôles
        role_ids = request.POST.getlist("role_ids")

        if not (nom and prenom and email):
            messages.error(request, "Nom, prénom et email obligatoires.")
            return redirect("gestion:responsable_update", pk=pk)

        if not role_ids:
            messages.error(request, "Veuillez sélectionner au moins un rôle.")
            return redirect("gestion:responsable_update", pk=pk)

        try:
            with connection.cursor() as cur:
                # Mettre à jour le responsable
                if new_pass:
                    cur.execute("""
                        UPDATE isms.responsable
                        SET nom=%s, prenom=%s, email=%s, is_admin=%s, mot_de_pass=%s
                        WHERE id_responsable=%s;
                    """, [nom, prenom, email, is_admin, new_pass, pk])
                else:
                    cur.execute("""
                        UPDATE isms.responsable
                        SET nom=%s, prenom=%s, email=%s, is_admin=%s
                        WHERE id_responsable=%s;
                    """, [nom, prenom, email, is_admin, pk])

                # ✅ Mettre à jour les rôles : supprimer les anciens + insérer les nouveaux
                cur.execute("DELETE FROM isms.responsable_role WHERE id_responsable=%s;", [pk])
                
                for role_id in role_ids:
                    cur.execute("""
                        INSERT INTO isms.responsable_role(id_responsable, id_role)
                        VALUES (%s, %s);
                    """, [pk, role_id])

            messages.success(request, "Responsable modifié avec ses rôles.")
        except IntegrityError as e:
            if 'email' in str(e).lower():
                messages.error(request, "Cet email est déjà utilisé.")
            else:
                messages.error(request, "Erreur : données invalides.")
        except DatabaseError as e:
            messages.error(request, f"Erreur base de données : {str(e)}")

        return redirect("gestion:responsable_list")

    return render(request, "gestion/responsables/update.html", {
        "resp": resp, 
        "roles": roles,
        "current_roles": current_roles
    })


@superadmin_required
def responsable_toggle_admin(request, pk):
    with connection.cursor() as cur:
        cur.execute("UPDATE isms.responsable SET is_admin = NOT is_admin WHERE id_responsable = %s;", [pk])
    messages.success(request, "Droit admin mis à jour.")
    return redirect("gestion:responsable_list")


@superadmin_required
def responsable_delete(request, pk):
    try:
        with connection.cursor() as cur:
            # La suppression en cascade supprimera aussi responsable_role
            cur.execute("DELETE FROM isms.responsable WHERE id_responsable = %s;", [pk])
        messages.success(request, "Responsable supprimé.")
    except IntegrityError:
        messages.error(request, "Impossible de supprimer : ce responsable est lié à d'autres données (ex: encadrement/jury).")
    return redirect("gestion:responsable_list")

# =========================
# UTILISATEURS: ETUDIANT
# =========================

@superadmin_required
def etudiant_list(request):
    q = (request.GET.get("q") or "").strip().lower()
    
    where = []
    params = []
    
    if q:
        like = f"%{q}%"
        where.append("""
            (LOWER(e.nom) LIKE %s 
             OR LOWER(e.prenom) LIKE %s 
             OR LOWER(e.email) LIKE %s)
        """)
        params.extend([like, like, like])
    
    sql = f"""
        SELECT 
            e.id_etudiant, 
            e.nom, 
            e.prenom, 
            e.email, 
            e.telephone, 
            e.niveau,
            e.id_departement,
            e.id_annee,
            d.nom_departement,
            a.libelle as annee_libelle
        FROM isms.etudiant e
        JOIN isms.departement d ON d.id_departement = e.id_departement
        JOIN isms.annee_universitaire a ON a.id_annee = e.id_annee
        {"WHERE " + " AND ".join(where) if where else ""}
        ORDER BY e.nom, e.prenom
    """
    
    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = dictfetchall(cur)
        
        # Récupérer les départements avec leurs niveaux
        cur.execute("""
            SELECT 
                d.id_departement, 
                d.nom_departement, 
                n.libelle as niveau_libelle
            FROM isms.departement d
            JOIN isms.niveau n ON n.id_niveau = d.id_niveau
            ORDER BY d.nom_departement
        """)
        departements = dictfetchall(cur)
        
        # Récupérer les années universitaires
        cur.execute("""
            SELECT id_annee, libelle, active
            FROM isms.annee_universitaire
            ORDER BY libelle DESC
        """)
        annees = dictfetchall(cur)
    
    return render(request, "gestion/etudiants/list.html", {
        "rows": rows,
        "q": request.GET.get("q", ""),
        "departements": departements,
        "annees": annees,
    })


@superadmin_required
def etudiant_create(request):
    if request.method == "POST":
        # Vérifier si c'est un upload en masse
        if 'excel_file' in request.FILES:
            return handle_bulk_student_upload(request)
        
        # Création individuelle d'un étudiant
        id_etudiant = (request.POST.get("id_etudiant") or "").strip()
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        telephone = (request.POST.get("telephone") or "").strip()
        niveau = (request.POST.get("niveau") or "").strip()
        mot_de_pass = (request.POST.get("mot_de_pass") or "").strip()
        id_departement = request.POST.get("id_departement")
        id_annee = request.POST.get("id_annee")

        # Validation des champs obligatoires (selon contraintes BD)
        if not all([id_etudiant, nom, prenom, email, mot_de_pass, id_departement, id_annee]):
            messages.error(request, "Les champs Matricule, Nom, Prénom, Email, Mot de passe, Département et Année sont obligatoires.")
            return redirect("gestion:etudiant_list")

        # Validation du format email
        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Format d'email invalide.")
            return redirect("gestion:etudiant_list")

        try:
            with connection.cursor() as cur:
                # Insertion selon la structure exacte de la table etudiant
                cur.execute("""
                    INSERT INTO isms.etudiant(
                        id_etudiant, nom, prenom, email, telephone, niveau, 
                        mot_de_pass, id_departement, id_annee
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    id_etudiant,
                    nom, 
                    prenom, 
                    email, 
                    telephone if telephone else None,  # NULL si vide
                    niveau if niveau else None,        # NULL si vide
                    mot_de_pass, 
                    id_departement, 
                    id_annee
                ])
            
            messages.success(request, f"Étudiant {id_etudiant} - {nom} {prenom} créé avec succès.")
        except IntegrityError as e:
            # Matricule ou email unique ou FK invalide
            error_str = str(e).lower()
            if 'id_etudiant' in error_str or 'primary key' in error_str:
                messages.error(request, f"Ce matricule ({id_etudiant}) est déjà utilisé par un autre étudiant.")
            elif 'email' in error_str:
                messages.error(request, f"Cet email ({email}) est déjà utilisé par un autre étudiant.")
            else:
                messages.error(request, "Erreur : Département ou Année invalide.")
        except DatabaseError as e:
            messages.error(request, f"Erreur base de données : {str(e)}")
        
        return redirect("gestion:etudiant_list")

    return redirect("gestion:etudiant_list")


def handle_bulk_student_upload(request):
    """Gestion de l'upload Excel pour création en masse d'étudiants"""
    excel_file = request.FILES.get('excel_file')
    
    if not excel_file:
        messages.error(request, "Aucun fichier sélectionné.")
        return redirect("gestion:etudiant_list")
    
    # Validation de l'extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, "Le fichier doit être au format Excel (.xlsx ou .xls).")
        return redirect("gestion:etudiant_list")
    
    try:
        # Charger le workbook
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Récupérer les départements et années pour validation
        with connection.cursor() as cur:
            # Départements : nom_departement (VARCHAR(80) UNIQUE)
            cur.execute("""
                SELECT d.id_departement, d.nom_departement
                FROM isms.departement d
                ORDER BY d.nom_departement
            """)
            # Créer dictionnaire: nom_departement.lower() -> id_departement
            departments = {row[1].strip().lower(): row[0] for row in cur.fetchall()}
            
            # Années universitaires : libelle (VARCHAR(20) UNIQUE)
            cur.execute("""
                SELECT id_annee, libelle 
                FROM isms.annee_universitaire
                ORDER BY libelle
            """)
            # Créer dictionnaire: libelle.lower() -> id_annee
            annees = {row[1].strip().lower(): row[0] for row in cur.fetchall()}
        
        # Parcourir les lignes (en sautant l'en-tête = ligne 1)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Ignorer les lignes vides
                if not row or all(cell is None for cell in row):
                    continue
                
                # Colonnes attendues dans l'Excel (SANS Filière):
                # [0]=Matricule (id_etudiant), [1]=Nom, [2]=Prénom, [3]=Email, [4]=Téléphone, 
                # [5]=Niveau, [6]=Mot de passe, [7]=Département, [8]=Année
                
                id_etudiant = str(row[0]).strip() if row[0] else ""
                nom = str(row[1]).strip() if row[1] else ""
                prenom = str(row[2]).strip() if row[2] else ""
                email = str(row[3]).strip().lower() if row[3] else ""
                telephone = str(row[4]).strip() if row[4] and str(row[4]).strip() else None
                niveau = str(row[5]).strip() if row[5] and str(row[5]).strip() else None
                mot_de_pass = str(row[6]).strip() if row[6] else ""
                departement_nom = str(row[7]).strip().lower() if row[7] else ""
                annee_libelle = str(row[8]).strip().lower() if row[8] else ""
                
                # Validation des champs obligatoires (selon contraintes BD)
                if not all([id_etudiant, nom, prenom, email, mot_de_pass, departement_nom, annee_libelle]):
                    errors.append(f"Ligne {idx}: Champs obligatoires manquants (Matricule, Nom, Prénom, Email, Mot de passe, Département, Année)")
                    error_count += 1
                    continue
                
                # Validation du format email (contrainte CHECK: position('@' in email) > 1)
                try:
                    validate_email(email)
                    if '@' not in email or email.index('@') == 0:
                        raise ValidationError("@ manquant ou mal placé")
                except ValidationError:
                    errors.append(f"Ligne {idx}: Email invalide ({email})")
                    error_count += 1
                    continue
                
                # Récupérer l'ID du département
                id_departement = departments.get(departement_nom)
                if not id_departement:
                    available_depts = ", ".join(list(departments.keys())[:5])
                    errors.append(f"Ligne {idx}: Département '{departement_nom}' introuvable. Disponibles: {available_depts}...")
                    error_count += 1
                    continue
                
                # Récupérer l'ID de l'année
                id_annee = annees.get(annee_libelle)
                if not id_annee:
                    available_annees = ", ".join(list(annees.keys()))
                    errors.append(f"Ligne {idx}: Année '{annee_libelle}' introuvable. Disponibles: {available_annees}")
                    error_count += 1
                    continue
                
                # Insertion dans la table etudiant (structure exacte de la BD)
                with connection.cursor() as cur:
                    cur.execute("""
                        INSERT INTO isms.etudiant(
                            id_etudiant, nom, prenom, email, telephone, niveau,
                            mot_de_pass, id_departement, id_annee
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        id_etudiant,
                        nom, 
                        prenom, 
                        email, 
                        telephone,  # NULL si None
                        niveau,     # NULL si None
                        mot_de_pass, 
                        id_departement, 
                        id_annee
                    ])
                
                success_count += 1
                
            except IntegrityError as e:
                # Contrainte UNIQUE sur id_etudiant (PRIMARY KEY) ou email
                error_str = str(e).lower()
                if 'id_etudiant' in error_str or 'primary key' in error_str:
                    errors.append(f"Ligne {idx}: Matricule déjà existant ({id_etudiant})")
                elif 'email' in error_str or 'unique' in error_str:
                    errors.append(f"Ligne {idx}: Email déjà existant ({email})")
                else:
                    errors.append(f"Ligne {idx}: Contrainte d'intégrité violée")
                error_count += 1
            except Exception as e:
                errors.append(f"Ligne {idx}: Erreur - {str(e)}")
                error_count += 1
        
        # Affichage des résultats
        if success_count > 0:
            messages.success(request, f"✅ {success_count} étudiant(s) importé(s) avec succès.")
        
        if error_count > 0:
            # Afficher les 10 premières erreurs
            error_list = "<ul class='mb-0 small'>" + "".join([f"<li>{e}</li>" for e in errors[:10]]) + "</ul>"
            if len(errors) > 10:
                error_list += f"<p class='mt-2 mb-0 small'><em>... et {len(errors) - 10} autre(s) erreur(s)</em></p>"
            messages.error(request, mark_safe(f"❌ {error_count} erreur(s) rencontrée(s): {error_list}"))
        
        if success_count == 0 and error_count == 0:
            messages.warning(request, "⚠️ Aucune donnée trouvée dans le fichier Excel (fichier vide ou uniquement en-tête).")
    
    except Exception as e:
        messages.error(request, f"❌ Erreur lors de la lecture du fichier Excel : {str(e)}")
    
    return redirect("gestion:etudiant_list")


@superadmin_required
def etudiant_template_download(request):
    """Générer et télécharger un modèle Excel pour l'import en masse"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etudiants"
    
    # En-têtes (SANS la colonne Filière)
    headers = [
        "Matricule", "Nom", "Prénom", "Email", "Téléphone",  
        "Niveau", "Mot de passe", "Département", "Année Universitaire"
    ]
    ws.append(headers)
    
    # Style des en-têtes
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Récupérer des exemples réels depuis la BD
    with connection.cursor() as cur:
        # Premier département
        cur.execute("""
            SELECT d.nom_departement
            FROM isms.departement d
            ORDER BY d.nom_departement
            LIMIT 1
        """)
        dept_example = cur.fetchone()
        dept_name = dept_example[0] if dept_example else "SEA"
        
        # Année active
        cur.execute("""
            SELECT libelle
            FROM isms.annee_universitaire
            WHERE active = TRUE
            LIMIT 1
        """)
        annee_example = cur.fetchone()
        if not annee_example:
            # Si pas d'année active, prendre la plus récente
            cur.execute("""
                SELECT libelle
                FROM isms.annee_universitaire
                ORDER BY id_annee DESC
                LIMIT 1
            """)
            annee_example = cur.fetchone()
        annee_name = annee_example[0] if annee_example else "2025-2026"
    
    # Ajouter des exemples de données (SANS Filière)
    ws.append([
        "22640", "Mohamed", "Mohamedou", "22640@isms.esp.mr", "+222 12345678",
        "SEA_L3", "22640pass", dept_name, annee_name
    ])
    ws.append([
        "23603", "Zeinebou Mohamed Lemine", "Ennahoui", "23603@isms.esp.mr", "+222 23456789",
        "SEA_L3", "23603pass", dept_name, annee_name
    ])
    
    # Créer un onglet d'instructions
    ws_instructions = wb.create_sheet("Instructions")
    
    # Récupérer les valeurs valides depuis la BD
    with connection.cursor() as cur:
        # Tous les départements
        cur.execute("""
            SELECT d.nom_departement, n.libelle as niveau
            FROM isms.departement d
            JOIN isms.niveau n ON n.id_niveau = d.id_niveau
            ORDER BY d.nom_departement
        """)
        depts = cur.fetchall()
        
        # Toutes les années
        cur.execute("""
            SELECT libelle, active
            FROM isms.annee_universitaire
            ORDER BY libelle DESC
        """)
        annees = cur.fetchall()
    
    # Contenu des instructions
    instructions = [
        ["📋 GUIDE D'IMPORT DES ÉTUDIANTS"],
        [""],
        ["1️⃣ COLONNES OBLIGATOIRES (ne peuvent pas être vides):"],
        ["   • Matricule (identifiant unique de l'étudiant, ex: 22640)"],
        ["   • Nom (max 60 caractères)"],
        ["   • Prénom (max 60 caractères)"],
        ["   • Email (max 120 caractères, doit contenir @, doit être unique)"],
        ["   • Mot de passe"],
        ["   • Département (doit correspondre EXACTEMENT à un département ci-dessous)"],
        ["   • Année Universitaire (doit correspondre EXACTEMENT à une année ci-dessous)"],
        [""],
        ["2️⃣ COLONNES OPTIONNELLES (peuvent être vides):"],
        ["   • Téléphone (max 30 caractères)"],
        ["   • Niveau (max 50 caractères, ex: SEA_L3, SEA_M1, etc.)"],
        [""],
        ["3️⃣ DÉPARTEMENTS DISPONIBLES DANS LA BASE:"],
    ]
    
    if depts:
        for dept in depts:
            instructions.append([f"   ✓ {dept[0]}", f"(Niveau associé: {dept[1]})"])
    else:
        instructions.append(["   ⚠️ Aucun département trouvé ! Contactez l'administrateur."])
    
    instructions.append([""])
    instructions.append(["4️⃣ ANNÉES UNIVERSITAIRES DISPONIBLES:"])
    
    if annees:
        for annee in annees:
            active_marker = " ⭐ ACTIVE" if annee[1] else ""
            instructions.append([f"   ✓ {annee[0]}{active_marker}"])
    else:
        instructions.append(["   ⚠️ Aucune année trouvée ! Contactez l'administrateur."])
    
    instructions.append([""])
    instructions.append(["⚠️ IMPORTANT:"])
    instructions.append(["   • Les MATRICULES doivent être UNIQUES (pas de doublons)"])
    instructions.append(["   • Les emails doivent être UNIQUES (pas de doublons)"])
    instructions.append(["   • Format conseillé pour les matricules: 22640, 23603, etc."])
    instructions.append(["   • Respectez EXACTEMENT les noms des départements et années"])
    instructions.append(["   • La casse (majuscules/minuscules) n'est PAS importante pour départements/années"])
    instructions.append(["   • Les espaces en début/fin sont automatiquement supprimés"])
    instructions.append(["   • Ligne vide = ligne ignorée"])
    
    # Ajouter les instructions
    for row_data in instructions:
        ws_instructions.append(row_data)
    
    # Style pour les instructions
    ws_instructions['A1'].font = Font(bold=True, size=16, color="0066CC")
    for row_idx in [3, 12, 16]:  # Titres de sections
        if row_idx <= ws_instructions.max_row:
            ws_instructions.cell(row_idx, 1).font = Font(bold=True, size=12)
    
    # Ajuster la largeur des colonnes
    for ws_current in [ws, ws_instructions]:
        for column in ws_current.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_current.column_dimensions[column_letter].width = min(max_length + 3, 60)
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_import_etudiants.xlsx'
    wb.save(response)
    
    return response


@superadmin_required
def etudiant_update(request, pk):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id_etudiant, nom, prenom, email
            FROM isms.etudiant
            WHERE id_etudiant=%s;
        """, [pk])
        row = cur.fetchone()

    if not row:
        messages.error(request, "Étudiant introuvable.")
        return redirect("gestion:etudiant_list")

    etu = {"id_etudiant": row[0], "nom": row[1], "prenom": row[2], "email": row[3]}

    if request.method == "POST":
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        mot_de_pass = (request.POST.get("mot_de_pass") or "").strip()  # optionnel

        if not (nom and prenom and email):
            messages.error(request, "Nom, prénom, email obligatoires.")
            return redirect("gestion:etudiant_update", pk=pk)

        with connection.cursor() as cur:
            if mot_de_pass:
                cur.execute("""
                    UPDATE isms.etudiant
                    SET nom=%s, prenom=%s, email=%s, mot_de_pass=%s
                    WHERE id_etudiant=%s;
                """, [nom, prenom, email, mot_de_pass, pk])
            else:
                cur.execute("""
                    UPDATE isms.etudiant
                    SET nom=%s, prenom=%s, email=%s
                    WHERE id_etudiant=%s;
                """, [nom, prenom, email, pk])

        messages.success(request, "Étudiant modifié.")
        return redirect("gestion:etudiant_list")

    return render(request, "gestion/etudiants/update.html", {"etu": etu})


@superadmin_required
def etudiant_delete(request, pk):
    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.etudiant WHERE id_etudiant=%s;", [pk])
        messages.success(request, "Étudiant supprimé.")
    except IntegrityError:
        messages.error(request, "Impossible de supprimer : étudiant lié à d'autres données (ex: mémoire, note...).")
    return redirect("gestion:etudiant_list")



# =========================
# METIER (PLACEHOLDERS)
# =========================
@superadmin_required
@admin_year_required
def memoire_list(request):
    annee_id = request.session.get("annee_id")

    q = (request.GET.get("q") or "").strip().lower()
    statut = (request.GET.get("statut") or "").strip()
    type_ = (request.GET.get("type") or "").strip()

    where = ["m.id_annee = %s"]
    params = [annee_id]

    if statut:
        where.append("m.statut = %s")
        params.append(statut)

    if type_:
        where.append("m.type = %s")
        params.append(type_)

    if q:
        like = f"%{q}%"
        where.append("""
          (LOWER(m.titre) LIKE %s
           OR LOWER(e.nom) LIKE %s
           OR LOWER(e.prenom) LIKE %s
           OR LOWER(e.email) LIKE %s)
        """)
        params.extend([like, like, like, like])

    where_sql = "WHERE " + " AND ".join(where)

    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT
              m.id_memoire, m.titre, m.type, m.statut, m.date_depot, m.fichier_pdf,
              e.id_etudiant, e.nom, e.prenom, e.email
            FROM isms.memoire m
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            {where_sql}
            ORDER BY m.id_memoire DESC;
        """, params)
        rows = fetchall_dict(cur)

        # étudiants de l'année active uniquement
        cur.execute("""
            SELECT id_etudiant, nom, prenom, email
            FROM isms.etudiant
            WHERE id_annee = %s
            ORDER BY id_etudiant DESC;
        """, [annee_id])
        etudiants = fetchall_dict(cur)

    statuts = ["DEPOSE", "EN_VERIFICATION", "VALIDE", "REFUSE"]
    types = ["PFE", "MEMOIRE", "RAPPORT", "THESE"]

    return render(request, "gestion/memoires/list.html", {
        "rows": rows,
        "etudiants": etudiants,
        "statuts": statuts,
        "types": types,
        "q": request.GET.get("q", ""),
        "statut_filtre": statut,
        "type_filtre": type_,
    })



def save_memoire_pdf(file_obj):
    folder = os.path.join(settings.MEDIA_ROOT, "memoires")
    os.makedirs(folder, exist_ok=True)

    ts = timezone.now().strftime("%Y%m%d_%H%M%S")
    safe_name = file_obj.name.replace(" ", "_")
    filename = f"{ts}_{safe_name}"

    abs_path = os.path.join(folder, filename)

    with open(abs_path, "wb+") as f:
        for chunk in file_obj.chunks():
            f.write(chunk)

    # IMPORTANT: on stocke en DB un chemin avec "/" (pas "\")
    rel_path = f"memoires/{filename}"
    return rel_path


@superadmin_required
@admin_year_required
def memoire_create(request):
    if request.method != "POST":
        return redirect("gestion:memoire_list")

    annee_id = request.session.get("annee_id")

    titre = (request.POST.get("titre") or "").strip()
    type_ = (request.POST.get("type") or "").strip()
    description = (request.POST.get("description") or "").strip()
    statut = (request.POST.get("statut") or "DEPOSE").strip()
    id_etudiant = request.POST.get("id_etudiant")

    if not (titre and type_ and id_etudiant):
        messages.error(request, "Titre, type et étudiant obligatoires.")
        return redirect("gestion:memoire_list")

    fichier_pdf = None
    if request.FILES.get("fichier_pdf"):
        fichier_pdf = save_memoire_pdf(request.FILES["fichier_pdf"])

    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO isms.memoire(titre, type, description, fichier_pdf, statut, id_etudiant, id_annee)
            VALUES (%s,%s,%s,%s,%s,%s,%s);
        """, [titre, type_, description or None, fichier_pdf, statut, id_etudiant, annee_id])

    messages.success(request, "Mémoire ajouté.")
    return redirect("gestion:memoire_list")

@superadmin_required
@admin_year_required
def memoire_update(request, pk):
    annee_id = request.session.get("annee_id")

    with connection.cursor() as cur:
        cur.execute("""
            SELECT id_memoire, titre, type, description, fichier_pdf, statut, id_etudiant
            FROM isms.memoire
            WHERE id_memoire=%s AND id_annee=%s;
        """, [pk, annee_id])
        row = cur.fetchone()

        cur.execute("""
            SELECT id_etudiant, nom, prenom, email
            FROM isms.etudiant
            WHERE id_annee=%s
            ORDER BY id_etudiant DESC;
        """, [annee_id])
        etudiants = fetchall_dict(cur)

    if not row:
        messages.error(request, "Mémoire introuvable (ou pas dans l'année active).")
        return redirect("gestion:memoire_list")

    mem = {
        "id_memoire": row[0],
        "titre": row[1],
        "type": row[2],
        "description": row[3],
        "fichier_pdf": row[4],
        "statut": row[5],
        "id_etudiant": row[6],
    }

    statuts = ["DEPOSE", "EN_VERIFICATION", "VALIDE", "REFUSE"]
    types = ["PFE", "MEMOIRE", "RAPPORT", "THESE"]

    if request.method == "POST":
        titre = (request.POST.get("titre") or "").strip()
        type_ = (request.POST.get("type") or "").strip()
        description = (request.POST.get("description") or "").strip()
        statut = (request.POST.get("statut") or "").strip()
        id_etudiant = request.POST.get("id_etudiant")

        if not (titre and type_ and statut and id_etudiant):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
            return redirect("gestion:memoire_update", pk=pk)

        set_parts = ["titre=%s", "type=%s", "description=%s", "statut=%s", "id_etudiant=%s"]
        params = [titre, type_, description or None, statut, id_etudiant]

        # nouveau PDF optionnel
        if request.FILES.get("fichier_pdf"):
            fichier_pdf = save_memoire_pdf(request.FILES["fichier_pdf"])
            set_parts.append("fichier_pdf=%s")
            params.append(fichier_pdf)

        params.extend([pk, annee_id])

        with connection.cursor() as cur:
            cur.execute(f"""
                UPDATE isms.memoire
                SET {", ".join(set_parts)}
                WHERE id_memoire=%s AND id_annee=%s;
            """, params)

        messages.success(request, "Mémoire modifié.")
        return redirect("gestion:memoire_list")

    return render(request, "gestion/memoires/update.html", {
        "mem": mem,
        "etudiants": etudiants,
        "statuts": statuts,
        "types": types,
    })


@superadmin_required
@admin_year_required
def memoire_delete(request, pk):
    annee_id = request.session.get("annee_id")
    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.memoire WHERE id_memoire=%s AND id_annee=%s;", [pk, annee_id])
        messages.success(request, "Mémoire supprimé.")
    except IntegrityError:
        messages.error(request, "Impossible de supprimer : mémoire lié à encadrement / soutenance / note.")
    return redirect("gestion:memoire_list")


@superadmin_required
@admin_year_required
def encadrement_list(request):
    annee_id = request.session.get("annee_id")

    q = (request.GET.get("q") or "").strip().lower()
    enc_type = (request.GET.get("encadrement") or "").strip()  # ENCADRANT / CO_ENCADRANT

    where = ["m.id_annee = %s"]
    params = [annee_id]

    if enc_type:
        where.append("en.encadrement = %s")
        params.append(enc_type)

    if q:
        like = f"%{q}%"
        where.append("""
          (
            LOWER(m.titre) LIKE %s
            OR LOWER(r.nom) LIKE %s OR LOWER(r.prenom) LIKE %s OR LOWER(r.email) LIKE %s
            OR LOWER(e.nom) LIKE %s OR LOWER(e.prenom) LIKE %s OR LOWER(e.email) LIKE %s
            OR LOWER(d.nom_departement) LIKE %s
          )
        """)
        params.extend([like, like, like, like, like, like, like, like])

    where_sql = "WHERE " + " AND ".join(where)

    with connection.cursor() as cur:
        # LISTE encadrements (avec étudiant + mémoire + dept)
        cur.execute(f"""
            SELECT
              en.id_responsable, en.id_memoire, en.encadrement,

              r.nom AS r_nom, r.prenom AS r_prenom, r.email AS r_email,

              m.titre AS m_titre, m.statut AS m_statut, m.type AS m_type,

              e.id_etudiant, e.nom AS e_nom, e.prenom AS e_prenom, e.email AS e_email,
              d.nom_departement AS dep_nom
            FROM isms.encadrement en
            JOIN isms.responsable r ON r.id_responsable = en.id_responsable
            JOIN isms.memoire m ON m.id_memoire = en.id_memoire
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            JOIN isms.departement d ON d.id_departement = e.id_departement
            {where_sql}
            ORDER BY m.id_memoire DESC, en.encadrement;
        """, params)
        rows = fetchall_dict(cur)

        # MEMOIRES de l'année active
        cur.execute("""
            SELECT m.id_memoire, m.titre, m.statut, m.type,
                   e.nom AS e_nom, e.prenom AS e_prenom
            FROM isms.memoire m
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            WHERE m.id_annee = %s
            ORDER BY m.id_memoire DESC;
        """, [annee_id])
        memoires = fetchall_dict(cur)

        # RESPONSABLES
        cur.execute("""
            SELECT id_responsable, nom, prenom, email
            FROM isms.responsable
            ORDER BY id_responsable DESC;
        """)
        responsables = fetchall_dict(cur)

    enc_types = ["ENCADRANT", "CO_ENCADRANT"]

    return render(request, "gestion/encadrements/list.html", {
        "rows": rows,
        "memoires": memoires,
        "responsables": responsables,
        "enc_types": enc_types,
        "q": request.GET.get("q", ""),
        "enc_filtre": enc_type,
    })


@superadmin_required
@admin_year_required
def encadrement_create(request):
    if request.method != "POST":
        return redirect("gestion:encadrement_list")

    annee_id = request.session.get("annee_id")

    id_responsable = request.POST.get("id_responsable")
    id_memoire = request.POST.get("id_memoire")
    enc_type = (request.POST.get("encadrement") or "").strip()

    if not (id_responsable and id_memoire and enc_type):
        messages.error(request, "Responsable, mémoire et type d'encadrement obligatoires.")
        return redirect("gestion:encadrement_list")

    # sécurité : mémoire doit être dans l'année active
    with connection.cursor() as cur:
        cur.execute("SELECT 1 FROM isms.memoire WHERE id_memoire=%s AND id_annee=%s;", [id_memoire, annee_id])
        if cur.fetchone() is None:
            messages.error(request, "Mémoire invalide (pas dans l'année active).")
            return redirect("gestion:encadrement_list")

    # règle : un seul ENCADRANT principal par mémoire
    if enc_type == "ENCADRANT":
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 1
                FROM isms.encadrement
                WHERE id_memoire=%s AND encadrement='ENCADRANT'
                LIMIT 1;
            """, [id_memoire])
            if cur.fetchone():
                messages.error(request, "Ce mémoire a déjà un encadrant principal.")
                return redirect("gestion:encadrement_list")

    try:
        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO isms.encadrement(id_responsable, id_memoire, encadrement)
                VALUES (%s,%s,%s);
            """, [id_responsable, id_memoire, enc_type])
        messages.success(request, "Encadrement ajouté.")
    except IntegrityError:
        messages.error(request, "Déjà existant : ce responsable est déjà lié à ce mémoire.")

    return redirect("gestion:encadrement_list")

    if request.method != "POST":
        return redirect("gestion:encadrement_list")

    annee_id = request.session.get("annee_id")

    id_responsable = request.POST.get("id_responsable")
    id_memoire = request.POST.get("id_memoire")
    enc_type = (request.POST.get("encadrement") or "").strip()

    if not (id_responsable and id_memoire and enc_type):
        messages.error(request, "Responsable, mémoire et type d'encadrement obligatoires.")
        return redirect("gestion:encadrement_list")

    # sécurité : le mémoire doit appartenir à l'année active
    with connection.cursor() as cur:
        cur.execute("SELECT 1 FROM isms.memoire WHERE id_memoire=%s AND id_annee=%s;", [id_memoire, annee_id])
        if cur.fetchone() is None:
            messages.error(request, "Mémoire invalide (pas dans l'année active).")
            return redirect("gestion:encadrement_list")

    try:
        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO isms.encadrement(id_responsable, id_memoire, encadrement)
                VALUES (%s,%s,%s);
            """, [id_responsable, id_memoire, enc_type])
        messages.success(request, "Encadrement ajouté.")
    except IntegrityError:
        messages.error(request, "Déjà existant : ce responsable est déjà lié à ce mémoire.")

    return redirect("gestion:encadrement_list")

@superadmin_required
@admin_year_required
def encadrement_update(request, id_responsable, id_memoire):
    annee_id = request.session.get("annee_id")

    enc_types = ["ENCADRANT", "CO_ENCADRANT"]

    with connection.cursor() as cur:
        # sécurité année (mémoire ancien doit appartenir à l'année active)
        cur.execute("SELECT 1 FROM isms.memoire WHERE id_memoire=%s AND id_annee=%s;", [id_memoire, annee_id])
        if cur.fetchone() is None:
            messages.error(request, "Action refusée (mémoire hors année active).")
            return redirect("gestion:encadrement_list")

        # ligne encadrement existante (old)
        cur.execute("""
            SELECT en.id_responsable, en.id_memoire, en.encadrement,
                   r.nom, r.prenom, r.email,
                   m.titre, m.statut, m.type,
                   e.nom, e.prenom, e.email
            FROM isms.encadrement en
            JOIN isms.responsable r ON r.id_responsable = en.id_responsable
            JOIN isms.memoire m ON m.id_memoire = en.id_memoire
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            WHERE en.id_responsable=%s AND en.id_memoire=%s;
        """, [id_responsable, id_memoire])
        row = cur.fetchone()

        # listes pour dropdowns
        cur.execute("""
            SELECT m.id_memoire, m.titre, m.statut, m.type,
                   e.nom AS e_nom, e.prenom AS e_prenom
            FROM isms.memoire m
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            WHERE m.id_annee=%s
            ORDER BY m.id_memoire DESC;
        """, [annee_id])
        memoires = fetchall_dict(cur)

        cur.execute("""
            SELECT id_responsable, nom, prenom, email
            FROM isms.responsable
            ORDER BY id_responsable DESC;
        """)
        responsables = fetchall_dict(cur)

    if not row:
        messages.error(request, "Encadrement introuvable.")
        return redirect("gestion:encadrement_list")

    # données actuelles (old)
    enc = {
        "old_id_responsable": row[0],
        "old_id_memoire": row[1],
        "old_encadrement": row[2],

        "r_nom": row[3], "r_prenom": row[4], "r_email": row[5],
        "m_titre": row[6], "m_statut": row[7], "m_type": row[8],
        "e_nom": row[9], "e_prenom": row[10], "e_email": row[11],
    }

    if request.method == "POST":
        new_id_responsable = request.POST.get("id_responsable")
        new_id_memoire = request.POST.get("id_memoire")
        new_enc_type = (request.POST.get("encadrement") or "").strip()

        if not (new_id_responsable and new_id_memoire and new_enc_type):
            messages.error(request, "Tous les champs sont obligatoires.")
            return redirect("gestion:encadrement_update", id_responsable=id_responsable, id_memoire=id_memoire)

        if new_enc_type not in enc_types:
            messages.error(request, "Type d'encadrement invalide.")
            return redirect("gestion:encadrement_update", id_responsable=id_responsable, id_memoire=id_memoire)

        # sécurité : mémoire choisi doit appartenir à l'année active
        with connection.cursor() as cur:
            cur.execute("SELECT 1 FROM isms.memoire WHERE id_memoire=%s AND id_annee=%s;", [new_id_memoire, annee_id])
            if cur.fetchone() is None:
                messages.error(request, "Mémoire invalide (pas dans l'année active).")
                return redirect("gestion:encadrement_update", id_responsable=id_responsable, id_memoire=id_memoire)

        # règle : un seul ENCADRANT principal par mémoire (nouveau mémoire)
        if new_enc_type == "ENCADRANT":
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT 1
                    FROM isms.encadrement
                    WHERE id_memoire=%s AND encadrement='ENCADRANT'
                      AND NOT (id_responsable=%s AND id_memoire=%s)
                    LIMIT 1;
                """, [new_id_memoire, enc["old_id_responsable"], enc["old_id_memoire"]])
                if cur.fetchone():
                    messages.error(request, "Ce mémoire a déjà un encadrant principal.")
                    return redirect("gestion:encadrement_update", id_responsable=id_responsable, id_memoire=id_memoire)

        # UPDATE "complet" = delete old + insert new (transaction)
        try:
            with connection.cursor() as cur:
                cur.execute("BEGIN;")

                # supprimer l'ancienne ligne
                cur.execute("""
                    DELETE FROM isms.encadrement
                    WHERE id_responsable=%s AND id_memoire=%s;
                """, [enc["old_id_responsable"], enc["old_id_memoire"]])

                # insérer la nouvelle ligne
                cur.execute("""
                    INSERT INTO isms.encadrement(id_responsable, id_memoire, encadrement)
                    VALUES (%s,%s,%s);
                """, [new_id_responsable, new_id_memoire, new_enc_type])

                cur.execute("COMMIT;")

            messages.success(request, "Encadrement modifié (responsable/mémoire/type).")
            return redirect("gestion:encadrement_list")

        except IntegrityError:
            # doublon PK ou contrainte
            with connection.cursor() as cur:
                cur.execute("ROLLBACK;")
            messages.error(request, "Impossible : cet encadrement existe déjà (doublon).")
            return redirect("gestion:encadrement_update", id_responsable=id_responsable, id_memoire=id_memoire)

    return render(request, "gestion/encadrements/update.html", {
        "enc": enc,
        "enc_types": enc_types,
        "memoires": memoires,
        "responsables": responsables,
        "old_id_responsable": enc["old_id_responsable"],
        "old_id_memoire": enc["old_id_memoire"],
    })

@superadmin_required
@admin_year_required
def encadrement_delete(request, id_responsable: int, id_memoire: int):
    annee_id = request.session.get("annee_id")

    try:
        with connection.cursor() as cur:
            # 1) vérifier que cet encadrement existe + récupérer son type
            cur.execute("""
                SELECT en.encadrement
                FROM isms.encadrement en
                JOIN isms.memoire m ON m.id_memoire = en.id_memoire
                WHERE en.id_responsable = %s
                  AND en.id_memoire = %s
                  AND m.id_annee = %s
                LIMIT 1
            """, [id_responsable, id_memoire, annee_id])
            row = cur.fetchone()

            if not row:
                messages.error(request, "Encadrement introuvable.")
                return redirect("gestion:encadrement_list")

            enc_type = row[0]  # 'ENCADRANT' ou 'CO_ENCADRANT'

            # 2) si c’est un ENCADRANT, vérifier combien d'encadrants principaux restent
            if enc_type == "ENCADRANT":
                cur.execute("""
                    SELECT COUNT(*)
                    FROM isms.encadrement en
                    WHERE en.id_memoire = %s AND en.encadrement = 'ENCADRANT'
                """, [id_memoire])
                nb_encadrants = cur.fetchone()[0]

                if nb_encadrants <= 1:
                    messages.error(
                        request,
                        "Impossible de supprimer : ce mémoire doit garder au moins un encadrant principal. "
                        "Ajoutez un autre encadrant puis réessayez."
                    )
                    return redirect("gestion:encadrement_list")

            # 3) suppression
            cur.execute("""
                DELETE FROM isms.encadrement
                WHERE id_responsable = %s AND id_memoire = %s
            """, [id_responsable, id_memoire])

        messages.success(request, "Encadrement supprimé.")
        return redirect("gestion:encadrement_list")

    except DatabaseError as e:
        msg = str(e)

        # message propre pour ton trigger
        if "fn_check_memoire_has_encadrant" in msg or "doit avoir au moins 1 encadrant" in msg:
            messages.error(
                request,
                "Suppression refusée : ce mémoire doit garder au moins un encadrant principal."
            )
            return redirect("gestion:encadrement_list")

        messages.error(request, "Erreur base de données.")
        return redirect("gestion:encadrement_list")
    annee_id = request.session.get("annee_id")

    # sécurité année
    with connection.cursor() as cur:
        cur.execute("SELECT 1 FROM isms.memoire WHERE id_memoire=%s AND id_annee=%s;", [id_memoire, annee_id])
        if cur.fetchone() is None:
            messages.error(request, "Action refusée (mémoire hors année active).")
            return redirect("gestion:encadrement_list")

    with connection.cursor() as cur:
        cur.execute("""
            DELETE FROM isms.encadrement
            WHERE id_responsable=%s AND id_memoire=%s;
        """, [id_responsable, id_memoire])

    messages.success(request, "Encadrement supprimé.")
    return redirect("gestion:encadrement_list")

############################ SOUTENANCE ##############################

def dictfetchall(cursor):
    cols = [col[0] for col in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


# =========================================================
# SOUTENANCE - LIST
# =========================================================
@superadmin_required
@admin_year_required
def soutenance_list(request):
    annee_id = request.session.get("annee_id")

    q = (request.GET.get("q") or "").strip().lower()
    statut = (request.GET.get("statut") or "").strip()

    where = ["s.id_annee = %s"]
    params = [annee_id]

    if statut:
        where.append("s.statut = %s")
        params.append(statut)

    if q:
        like = f"%{q}%"
        where.append("""
            (LOWER(m.titre) LIKE %s
             OR LOWER(j.nom_jury) LIKE %s
             OR LOWER(sa.nom_salle) LIKE %s
             OR LOWER(e.nom) LIKE %s
             OR LOWER(e.prenom) LIKE %s
             OR LOWER(e.email) LIKE %s)
        """)
        params.extend([like, like, like, like, like, like])

    sql = f"""
        SELECT
            s.id_soutenance, s.date_, s.heure, s.statut,
            sa.nom_salle,
            j.nom_jury,
            m.id_memoire, m.titre,
            e.nom AS etu_nom, e.prenom AS etu_prenom, e.email AS etu_email
        FROM isms.soutenance s
        JOIN isms.memoire m ON m.id_memoire = s.id_memoire
        JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
        JOIN isms.jury j ON j.id_jury = s.id_jury
        JOIN isms.salle sa ON sa.id_salle = s.id_salle
        WHERE {" AND ".join(where)}
        ORDER BY s.date_ DESC, s.heure DESC, s.id_soutenance DESC
    """

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = dictfetchall(cur)

    return render(request, "gestion/soutenances/list.html", {
        "rows": rows,
        "q": request.GET.get("q", ""),
        "statut": statut,
        "STATUTS": ["PLANIFIEE", "EFFECTUEE", "ANNULEE"],
    })


# =========================================================
# SOUTENANCE - CREATE
# =========================================================

@superadmin_required
@admin_year_required
def soutenance_create(request):
    annee_id = request.session.get("annee_id")

    # ---------------------------
    # Helpers: messages DB propres
    # ---------------------------
    def _db_message(e: Exception) -> str:
        raw = str(e)

        # 1) Trigger: mémoire non VALIDE
        if "fn_soutenance_requires_memoire_valide" in raw or "mémoire non VALIDE" in raw:
            # Extraire statut si présent: statut=DEPOSE
            m = re.search(r"statut=([A-Z_]+)", raw)
            st = m.group(1) if m else "INCONNU"
            return f"Impossible de planifier : le mémoire n'est pas validé (statut={st})."

        # 2) Trigger: conflit encadrant/co-encadrant dans jury
        if "Conflit: un encadrant du mémoire" in raw or "fn_" in raw and "encadrant" in raw.lower():
            return "Conflit : un encadrant/co-encadrant du mémoire ne peut pas faire partie du jury choisi."

        # 3) Unique id_memoire dans soutenance => déjà planifié
        # (selon le message postgres, ça peut contenir 'unique' ou le nom du constraint)
        if "duplicate key value violates unique constraint" in raw and "id_memoire" in raw:
            return "Impossible : ce mémoire a déjà une soutenance planifiée."

        # 4) Fallback propre
        return "Erreur base de données. Veuillez vérifier les informations saisies."

    # ---------------------------
    # Data pour le formulaire
    # ---------------------------
    with connection.cursor() as cur:
        # ✅ IMPORTANT : ne proposer que les mémoires VALIDE et sans soutenance
        cur.execute("""
            SELECT m.id_memoire, m.titre, e.nom, e.prenom
            FROM isms.memoire m
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            WHERE m.id_annee = %s
              AND m.statut = 'VALIDE'
              AND NOT EXISTS (
                  SELECT 1 FROM isms.soutenance s WHERE s.id_memoire = m.id_memoire
              )
            ORDER BY m.id_memoire DESC
        """, [annee_id])
        memoires = dictfetchall(cur)

        cur.execute("SELECT id_jury, nom_jury FROM isms.jury ORDER BY nom_jury")
        jurys = dictfetchall(cur)

        cur.execute("SELECT id_salle, nom_salle FROM isms.salle ORDER BY nom_salle")
        salles = dictfetchall(cur)

    # ---------------------------
    # POST
    # ---------------------------
    if request.method == "POST":
        data = request.POST

        date_ = (data.get("date_") or "").strip()
        heure = (data.get("heure") or "").strip()
        statut = (data.get("statut") or "PLANIFIEE").strip()
        id_memoire = (data.get("id_memoire") or "").strip()
        id_jury = (data.get("id_jury") or "").strip()
        id_salle = (data.get("id_salle") or "").strip()

        errors = []

        # Validation simple champs requis
        if not date_:
            errors.append("La date est obligatoire.")
        if not heure:
            errors.append("L'heure est obligatoire.")
        if not id_memoire:
            errors.append("Le mémoire est obligatoire.")
        if not id_jury:
            errors.append("Le jury est obligatoire.")
        if not id_salle:
            errors.append("La salle est obligatoire.")

        if statut not in ("PLANIFIEE", "EFFECTUEE", "ANNULEE"):
            errors.append("Statut invalide.")

        # Si erreurs -> afficher toutes les erreurs
        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            try:
                # Optionnel mais top UX: re-vérifier côté serveur que le mémoire est VALIDE
                with connection.cursor() as cur:
                    cur.execute("""
                        SELECT statut
                        FROM isms.memoire
                        WHERE id_memoire=%s AND id_annee=%s
                        LIMIT 1
                    """, [id_memoire, annee_id])
                    row = cur.fetchone()

                if not row:
                    messages.error(request, "Mémoire introuvable.")
                elif row[0] != "VALIDE":
                    messages.error(request, f"Impossible de planifier : mémoire non validé (statut={row[0]}).")
                else:
                    with connection.cursor() as cur:
                        cur.execute("""
                            INSERT INTO isms.soutenance (date_, heure, statut, id_memoire, id_jury, id_annee, id_salle)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, [date_, heure, statut, id_memoire, id_jury, annee_id, id_salle])

                    messages.success(request, "Soutenance créée avec succès.")
                    return redirect("gestion:soutenance_list")

            except IntegrityError as e:
                # On donne un message plus précis si possible
                messages.error(request, _db_message(e))

            except DatabaseError as e:
                messages.error(request, _db_message(e))
                # Option debug : affiche détail brut uniquement en DEBUG
                if getattr(settings, "DEBUG", False):
                    messages.info(request, f"(DEBUG) {str(e)}")

    return render(request, "gestion/soutenances/form.html", {
        "mode": "create",
        "memoires": memoires,
        "jurys": jurys,
        "salles": salles,
        "STATUTS": ["PLANIFIEE", "EFFECTUEE", "ANNULEE"],
        "data": request.POST if request.method == "POST" else {},
    })

# =========================================================
# SOUTENANCE - UPDATE
# =========================================================
@superadmin_required
@admin_year_required
def soutenance_update(request, pk):
    annee_id = request.session.get("annee_id")

    with connection.cursor() as cur:
        cur.execute("""
            SELECT id_soutenance, date_, heure, statut, id_memoire, id_jury, id_salle
            FROM isms.soutenance
            WHERE id_soutenance=%s AND id_annee=%s
        """, [pk, annee_id])
        rows = dictfetchall(cur)

    if not rows:
        messages.error(request, "Soutenance introuvable.")
        return redirect("gestion:soutenance_list")

    data = rows[0]

    with connection.cursor() as cur:
        # Pour update: on affiche les mémoires de l'année,
        # mais on doit éviter d'affecter un mémoire déjà pris par une autre soutenance
        cur.execute("""
            SELECT m.id_memoire, m.titre, e.nom, e.prenom
            FROM isms.memoire m
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            WHERE m.id_annee=%s
              AND (
                  m.id_memoire = %s
                  OR NOT EXISTS (SELECT 1 FROM isms.soutenance s WHERE s.id_memoire = m.id_memoire)
              )
            ORDER BY m.id_memoire DESC
        """, [annee_id, data["id_memoire"]])
        memoires = dictfetchall(cur)

        cur.execute("SELECT id_jury, nom_jury FROM isms.jury ORDER BY nom_jury")
        jurys = dictfetchall(cur)

        cur.execute("SELECT id_salle, nom_salle FROM isms.salle ORDER BY nom_salle")
        salles = dictfetchall(cur)

    if request.method == "POST":
        date_ = request.POST.get("date_")
        heure = request.POST.get("heure")
        statut = (request.POST.get("statut") or "").strip()
        id_memoire = request.POST.get("id_memoire")
        id_jury = request.POST.get("id_jury")
        id_salle = request.POST.get("id_salle")

        if not date_ or not heure or not statut or not id_memoire or not id_jury or not id_salle:
            messages.error(request, "Tous les champs sont obligatoires.")
        else:
            try:
                with connection.cursor() as cur:
                    cur.execute("""
                        UPDATE isms.soutenance
                        SET date_=%s, heure=%s, statut=%s,
                            id_memoire=%s, id_jury=%s, id_salle=%s
                        WHERE id_soutenance=%s AND id_annee=%s
                    """, [date_, heure, statut, id_memoire, id_jury, id_salle, pk, annee_id])
                messages.success(request, "Soutenance modifiée.")
                return redirect("gestion:soutenance_list")
            except IntegrityError:
                messages.error(request, "Modification impossible : mémoire déjà planifié / contrainte DB.")

            except DatabaseError as e:
                msg = str(e)
                if "Conflit: un encadrant du mémoire" in msg:
                    messages.error(request, "Conflit : un encadrant (ou co-encadrant) de ce mémoire ne peut pas être membre du jury choisi.")
                else:
                    messages.error(request, "Erreur base de données : " + msg)

        data = {
            "date_": date_, "heure": heure, "statut": statut,
            "id_memoire": int(id_memoire), "id_jury": int(id_jury), "id_salle": int(id_salle)
        }

    return render(request, "gestion/soutenances/form.html", {
        "mode": "update",
        "pk": pk,
        "memoires": memoires,
        "jurys": jurys,
        "salles": salles,
        "STATUTS": ["PLANIFIEE", "EFFECTUEE", "ANNULEE"],
        "data": data,
    })


# =========================================================
# SOUTENANCE - DELETE
# =========================================================
@superadmin_required
@admin_year_required
def soutenance_delete(request, pk):
    annee_id = request.session.get("annee_id")

    with connection.cursor() as cur:
        cur.execute("""
            SELECT s.id_soutenance, s.date_, s.heure, m.titre
            FROM isms.soutenance s
            JOIN isms.memoire m ON m.id_memoire = s.id_memoire
            WHERE s.id_soutenance=%s AND s.id_annee=%s
        """, [pk, annee_id])
        rows = dictfetchall(cur)

    if not rows:
        messages.error(request, "Soutenance introuvable.")
        return redirect("gestion:soutenance_list")

    row = rows[0]

    if request.method == "POST":
        try:
            with connection.cursor() as cur:
                cur.execute("""
                    DELETE FROM isms.soutenance
                    WHERE id_soutenance=%s AND id_annee=%s
                """, [pk, annee_id])
            messages.success(request, "Soutenance supprimée.")
            return redirect("gestion:soutenance_list")
        except IntegrityError:
            messages.error(request, "Suppression impossible (contrainte DB).")
        
        except DatabaseError as e:
            messages.error(request, "Erreur base de données : " + str(e))


    return render(request, "gestion/soutenances/delete.html", {"row": row})

# =========================================================
# JURY - LIST
# =========================================================
@superadmin_required
def jury_list(request):
    q = (request.GET.get("q") or "").strip().lower()

    where = []
    params = []

    if q:
        like = f"%{q}%"
        where.append("""
            (LOWER(j.nom_jury) LIKE %s
             OR EXISTS (
                 SELECT 1
                 FROM isms.composition_jury cj
                 JOIN isms.responsable r ON r.id_responsable = cj.id_responsable
                 WHERE cj.id_jury = j.id_jury
                   AND (LOWER(r.nom) LIKE %s OR LOWER(r.prenom) LIKE %s OR LOWER(r.email) LIKE %s)
             )
            )
        """)
        params.extend([like, like, like, like])

    sql = f"""
        SELECT
            j.id_jury,
            j.nom_jury,

            -- nombre de membres
            (SELECT COUNT(*)
             FROM isms.composition_jury cj
             WHERE cj.id_jury = j.id_jury) AS nb_membres,

            -- preview des membres (3 noms max)
            (SELECT STRING_AGG(x.full_name, ', ')
             FROM (
                SELECT (r.nom || ' ' || r.prenom) AS full_name
                FROM isms.composition_jury cj2
                JOIN isms.responsable r ON r.id_responsable = cj2.id_responsable
                WHERE cj2.id_jury = j.id_jury
                ORDER BY r.nom, r.prenom
                LIMIT 3
             ) x
            ) AS membres_preview

        FROM isms.jury j
        {"WHERE " + " AND ".join(where) if where else ""}
        ORDER BY j.nom_jury
    """

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = dictfetchall(cur)

    # petit affichage propre si aucun membre
    for r in rows:
        if r.get("nb_membres", 0) == 0:
            r["membres_preview"] = "Aucun membre"
        elif r.get("nb_membres", 0) > 3:
            r["membres_preview"] = (r["membres_preview"] or "") + " …"

    return render(request, "gestion/jurys/list.html", {
        "rows": rows,
        "q": request.GET.get("q", ""),
    })

@superadmin_required
def jury_create(request):

    # Responsables disponibles (excluant ceux déjà dans d'autres jurys)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT
                r.id_responsable, 
                r.nom, 
                r.prenom, 
                r.email,
                STRING_AGG(ro.code, ', ' ORDER BY ro.code) AS role_code
            FROM isms.responsable r
            LEFT JOIN isms.responsable_role rr ON rr.id_responsable = r.id_responsable
            LEFT JOIN isms.role ro ON ro.id_role = rr.id_role
            LEFT JOIN isms.composition_jury cj ON cj.id_responsable = r.id_responsable
            WHERE cj.id_responsable IS NULL
            GROUP BY r.id_responsable, r.nom, r.prenom, r.email
            ORDER BY r.nom, r.prenom
        """)
        responsables = dictfetchall(cur)

    if request.method == "POST":
        nom_jury = (request.POST.get("nom_jury") or "").strip()
        membres_ids = request.POST.getlist("membres")

        if not nom_jury:
            messages.error(request, "Le nom du jury est obligatoire.")
        elif not membres_ids:
            messages.error(request, "Veuillez sélectionner au moins un membre du jury.")
        else:
            try:
                with connection.cursor() as cur:
                    # Double vérification: s'assurer qu'aucun membre n'est dans un autre jury
                    placeholders = ','.join(['%s'] * len(membres_ids))
                    cur.execute(f"""
                        SELECT 
                            r.id_responsable,
                            r.nom,
                            r.prenom,
                            j.nom_jury
                        FROM isms.composition_jury cj
                        JOIN isms.responsable r ON r.id_responsable = cj.id_responsable
                        JOIN isms.jury j ON j.id_jury = cj.id_jury
                        WHERE cj.id_responsable IN ({placeholders})
                    """, membres_ids)
                    
                    membres_occupes = cur.fetchall()
                    
                    if membres_occupes:
                        conflits = []
                        for membre in membres_occupes:
                            conflits.append(f"{membre[1]} {membre[2]} (déjà membre du jury '{membre[3]}')")
                        
                        messages.error(
                            request, 
                            f"Les membres suivants ont été assignés à d'autres jurys entre-temps : {', '.join(conflits)}"
                        )
                    else:
                        # Tous les membres sont disponibles, créer le jury
                        cur.execute("""
                            INSERT INTO isms.jury (nom_jury)
                            VALUES (%s)
                            RETURNING id_jury
                        """, [nom_jury])
                        jury_id = cur.fetchone()[0]

                        # Insérer les membres
                        for rid in membres_ids:
                            cur.execute("""
                                INSERT INTO isms.composition_jury (id_responsable, id_jury)
                                VALUES (%s, %s)
                            """, [rid, jury_id])

                        messages.success(request, "Jury créé avec succès avec ses membres.")
                        return redirect("gestion:composition_jury_list", jury_id=jury_id)

            except IntegrityError as e:
                messages.error(request, "Erreur : données invalides ou doublon.")

            except DatabaseError as e:
                messages.error(request, "Erreur base de données : " + str(e))

    return render(request, "gestion/jurys/form.html", {
        "mode": "create",
        "responsables": responsables,
        "data": request.POST if request.method == "POST" else {},
        "selected_membres": request.POST.getlist("membres") if request.method == "POST" else [],
    })

# =========================================================
# JURY - UPDATE
# =========================================================
@superadmin_required
def jury_update(request, pk):
    with connection.cursor() as cur:
        cur.execute("SELECT id_jury, nom_jury FROM isms.jury WHERE id_jury=%s", [pk])
        rows = dictfetchall(cur)

    if not rows:
        messages.error(request, "Jury introuvable.")
        return redirect("gestion:jury_list")

    data = rows[0]

    # Récupérer les responsables disponibles (ceux pas dans d'autres jurys OU déjà dans CE jury)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT
                r.id_responsable, 
                r.nom, 
                r.prenom, 
                r.email,
                STRING_AGG(ro.code, ', ' ORDER BY ro.code) AS role_code
            FROM isms.responsable r
            LEFT JOIN isms.responsable_role rr ON rr.id_responsable = r.id_responsable
            LEFT JOIN isms.role ro ON ro.id_role = rr.id_role
            WHERE NOT EXISTS (
                SELECT 1 FROM isms.composition_jury cj
                WHERE cj.id_responsable = r.id_responsable
                AND cj.id_jury != %s
            )
            GROUP BY r.id_responsable, r.nom, r.prenom, r.email
            ORDER BY r.nom, r.prenom
        """, [pk])
        responsables = dictfetchall(cur)
        
        # Récupérer les membres actuels du jury
        cur.execute("""
            SELECT id_responsable
            FROM isms.composition_jury
            WHERE id_jury = %s
        """, [pk])
        current_membres = [str(row[0]) for row in cur.fetchall()]

    if request.method == "POST":
        nom_jury = (request.POST.get("nom_jury") or "").strip()
        membres_ids = request.POST.getlist("membres")

        if not nom_jury:
            messages.error(request, "Le nom du jury est obligatoire.")
        elif not membres_ids:
            messages.error(request, "Veuillez sélectionner au moins un membre du jury.")
        else:
            try:
                with connection.cursor() as cur:
                    # Vérifier si les nouveaux membres sont déjà dans d'autres jurys
                    nouveaux_membres = [m for m in membres_ids if m not in current_membres]
                    
                    if nouveaux_membres:
                        placeholders = ','.join(['%s'] * len(nouveaux_membres))
                        cur.execute(f"""
                            SELECT 
                                r.id_responsable,
                                r.nom,
                                r.prenom,
                                j.nom_jury
                            FROM isms.composition_jury cj
                            JOIN isms.responsable r ON r.id_responsable = cj.id_responsable
                            JOIN isms.jury j ON j.id_jury = cj.id_jury
                            WHERE cj.id_responsable IN ({placeholders})
                            AND cj.id_jury != %s
                        """, nouveaux_membres + [pk])
                        
                        membres_occupes = cur.fetchall()
                        
                        if membres_occupes:
                            conflits = []
                            for membre in membres_occupes:
                                conflits.append(f"{membre[1]} {membre[2]} (déjà membre du jury '{membre[3]}')")
                            
                            messages.error(
                                request, 
                                f"Les membres suivants sont déjà assignés à d'autres jurys : {', '.join(conflits)}"
                            )
                            data = {"id_jury": pk, "nom_jury": nom_jury}
                            return render(request, "gestion/jurys/form.html", {
                                "mode": "update",
                                "pk": pk,
                                "data": data,
                                "responsables": responsables,
                                "selected_membres": membres_ids,
                            })
                    
                    # Tout est OK, mettre à jour le jury
                    cur.execute("""
                        UPDATE isms.jury
                        SET nom_jury=%s
                        WHERE id_jury=%s
                    """, [nom_jury, pk])
                    
                    # Supprimer les anciennes compositions
                    cur.execute("""
                        DELETE FROM isms.composition_jury
                        WHERE id_jury = %s
                    """, [pk])
                    
                    # Insérer les nouvelles compositions
                    for rid in membres_ids:
                        cur.execute("""
                            INSERT INTO isms.composition_jury (id_responsable, id_jury)
                            VALUES (%s, %s)
                        """, [rid, pk])
                    
                messages.success(request, "Jury modifié avec succès.")
                return redirect("gestion:jury_list")

            except IntegrityError:
                messages.error(request, "Erreur : ce nom de jury est déjà utilisé ou données invalides.")

            except DatabaseError as e:
                messages.error(request, "Erreur base de données : " + str(e))

        data = {"id_jury": pk, "nom_jury": nom_jury}

    return render(request, "gestion/jurys/form.html", {
        "mode": "update",
        "pk": pk,
        "data": data,
        "responsables": responsables,
        "selected_membres": current_membres if request.method == "GET" else membres_ids,
    })

# =========================================================
# JURY - DELETE
# =========================================================
@superadmin_required
def jury_delete(request, pk):
    with connection.cursor() as cur:
        cur.execute("SELECT id_jury, nom_jury FROM isms.jury WHERE id_jury=%s", [pk])
        rows = dictfetchall(cur)

    if not rows:
        messages.error(request, "Jury introuvable.")
        return redirect("gestion:jury_list")

    row = rows[0]

    if request.method == "POST":
        try:
            with connection.cursor() as cur:
                cur.execute("DELETE FROM isms.jury WHERE id_jury=%s", [pk])
            messages.success(request, "Jury supprimé.")
            return redirect("gestion:jury_list")

        except IntegrityError:
            messages.error(request, "Suppression impossible : ce jury est utilisé (soutenance existante).")

        except DatabaseError as e:
            messages.error(request, "Erreur base de données : " + str(e))

    return render(request, "gestion/jurys/delete.html", {"row": row})

# =========================================================
# COMPOSITION JURY - LIST (membres d'un jury)
# =========================================================
@superadmin_required
def composition_jury_list(request, jury_id):
    q = (request.GET.get("q") or "").strip().lower()

    # Jury info
    with connection.cursor() as cur:
        cur.execute("SELECT id_jury, nom_jury FROM isms.jury WHERE id_jury=%s", [jury_id])
        jury_rows = dictfetchall(cur)
    if not jury_rows:
        messages.error(request, "Jury introuvable.")
        return redirect("gestion:jury_list")

    jury = jury_rows[0]

    # Membres actuels du jury
    where = ["cj.id_jury = %s"]
    params = [jury_id]

    if q:
        like = f"%{q}%"
        where.append("""
            (LOWER(r.nom) LIKE %s
             OR LOWER(r.prenom) LIKE %s
             OR LOWER(r.email) LIKE %s
             OR LOWER(roles_str) LIKE %s)
        """)
        params.extend([like, like, like, like])

    sql_membres = f"""
        SELECT
            r.id_responsable, 
            r.nom, 
            r.prenom, 
            r.email,
            STRING_AGG(ro.code, ', ' ORDER BY ro.code) AS role_code,
            STRING_AGG(ro.libelle, ', ' ORDER BY ro.libelle) AS role_libelle,
            STRING_AGG(ro.code || ' - ' || ro.libelle, ', ' ORDER BY ro.code) AS roles_str
        FROM isms.composition_jury cj
        JOIN isms.responsable r ON r.id_responsable = cj.id_responsable
        LEFT JOIN isms.responsable_role rr ON rr.id_responsable = r.id_responsable
        LEFT JOIN isms.role ro ON ro.id_role = rr.id_role
        WHERE {" AND ".join(where)}
        GROUP BY r.id_responsable, r.nom, r.prenom, r.email, cj.id_jury
        ORDER BY r.nom, r.prenom
    """

    with connection.cursor() as cur:
        cur.execute(sql_membres, params)
        membres = dictfetchall(cur)

    # Responsables ajoutables (pas dans AUCUN jury)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT
                r.id_responsable, 
                r.nom, 
                r.prenom, 
                r.email,
                STRING_AGG(ro.code, ', ' ORDER BY ro.code) AS role_code
            FROM isms.responsable r
            LEFT JOIN isms.responsable_role rr ON rr.id_responsable = r.id_responsable
            LEFT JOIN isms.role ro ON ro.id_role = rr.id_role
            LEFT JOIN isms.composition_jury cj ON cj.id_responsable = r.id_responsable
            WHERE cj.id_responsable IS NULL
            GROUP BY r.id_responsable, r.nom, r.prenom, r.email
            ORDER BY r.nom, r.prenom
        """)
        candidates = dictfetchall(cur)

    return render(request, "gestion/jurys/membres_list.html", {
        "jury": jury,
        "membres": membres,
        "candidates": candidates,
        "q": request.GET.get("q", ""),
    })


# =========================================================
# COMPOSITION JURY - ADD
# =========================================================
@superadmin_required
def composition_jury_add(request, jury_id):
    if request.method != "POST":
        return redirect("gestion:composition_jury_list", jury_id=jury_id)

    responsable_id = request.POST.get("id_responsable")
    if not responsable_id:
        messages.error(request, "Veuillez sélectionner un membre.")
        return redirect("gestion:composition_jury_list", jury_id=jury_id)

    try:
        with connection.cursor() as cur:
            # Vérifier que le responsable n'est pas déjà dans un autre jury
            cur.execute("""
                SELECT j.nom_jury
                FROM isms.composition_jury cj
                JOIN isms.jury j ON j.id_jury = cj.id_jury
                WHERE cj.id_responsable = %s
            """, [responsable_id])
            existing = cur.fetchone()
            
            if existing:
                messages.error(request, f"Ce membre est déjà assigné au jury '{existing[0]}'.")
                return redirect("gestion:composition_jury_list", jury_id=jury_id)
            
            # Ajouter le membre
            cur.execute("""
                INSERT INTO isms.composition_jury (id_responsable, id_jury)
                VALUES (%s, %s)
            """, [responsable_id, jury_id])
        messages.success(request, "Membre ajouté au jury.")
    except IntegrityError:
        messages.error(request, "Ce membre est déjà dans ce jury (ou contrainte DB).")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:composition_jury_list", jury_id=jury_id)


# =========================================================
# COMPOSITION JURY - REMOVE
# =========================================================
@superadmin_required
def composition_jury_remove(request, jury_id, responsable_id):
    if request.method != "POST":
        return redirect("gestion:composition_jury_list", jury_id=jury_id)

    try:
        with connection.cursor() as cur:
            cur.execute("""
                DELETE FROM isms.composition_jury
                WHERE id_jury=%s AND id_responsable=%s
            """, [jury_id, responsable_id])
        messages.success(request, "Membre retiré du jury.")
    except IntegrityError:
        messages.error(request, "Suppression impossible (contrainte DB).")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:composition_jury_list", jury_id=jury_id)

# =========================================================
# COMPOSITION JURY - CLEAR (supprimer tous les membres)
# =========================================================
@superadmin_required
def composition_jury_clear(request, jury_id):
    if request.method != "POST":
        return redirect("gestion:composition_jury_list", jury_id=jury_id)

    try:
        with connection.cursor() as cur:
            cur.execute("DELETE FROM isms.composition_jury WHERE id_jury=%s", [jury_id])
        messages.success(request, "Tous les membres ont été retirés de ce jury.")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:composition_jury_list", jury_id=jury_id)

############################## Notes ##############################

@superadmin_required
@admin_year_required
def note_list(request):
    annee_id = request.session.get("annee_id")

    q = (request.GET.get("q") or "").strip().lower()
    statut = (request.GET.get("statut") or "").strip()

    where = ["s.id_annee = %s"]
    params = [annee_id]

    if statut:
        where.append("s.statut = %s")
        params.append(statut)

    if q:
        like = f"%{q}%"
        where.append("""
            (CAST(e.id_etudiant AS TEXT) LIKE %s
             OR LOWER(e.nom) LIKE %s
             OR LOWER(e.prenom) LIKE %s
             OR LOWER(e.email) LIKE %s
             OR LOWER(m.titre) LIKE %s
             OR LOWER(j.nom_jury) LIKE %s)
        """)
        params.extend([like, like, like, like, like, like])

    sql = f"""
        SELECT
            s.id_soutenance,
            s.date_, s.heure, s.statut,
            e.id_etudiant, e.nom AS etu_nom, e.prenom AS etu_prenom, e.email AS etu_email,
            m.id_memoire, m.titre,
            j.nom_jury,
            sa.nom_salle,

            (SELECT ROUND(AVG(nj.note)::numeric, 2)
             FROM isms.note_jury nj
             WHERE nj.id_soutenance = s.id_soutenance) AS moyenne_harmonisee,

            n.id_note, n.note_finale, n.commentaire AS commentaire_final
        FROM isms.soutenance s
        JOIN isms.memoire m ON m.id_memoire = s.id_memoire
        JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
        JOIN isms.jury j ON j.id_jury = s.id_jury
        JOIN isms.salle sa ON sa.id_salle = s.id_salle
        LEFT JOIN isms.note n ON n.id_soutenance = s.id_soutenance
        WHERE {" AND ".join(where)}
        ORDER BY s.date_ DESC, s.heure DESC, s.id_soutenance DESC
    """

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = dictfetchall(cur)

    return render(request, "gestion/notes/list.html", {
        "rows": rows,
        "q": request.GET.get("q", ""),
        "statut": statut,
        "STATUTS": ["PLANIFIEE", "EFFECTUEE", "ANNULEE"],
    })

@superadmin_required
@admin_year_required
def note_detail(request, soutenance_id):
    annee_id = request.session.get("annee_id")

    # Base infos
    with connection.cursor() as cur:
        cur.execute("""
            SELECT
                s.id_soutenance, s.date_, s.heure, s.statut,
                sa.nom_salle,
                j.id_jury, j.nom_jury,

                m.id_memoire, m.titre, m.type, m.statut AS mem_statut, m.date_depot, m.description,
                e.id_etudiant, e.nom AS etu_nom, e.prenom AS etu_prenom, e.email AS etu_email, e.telephone, e.niveau,

                n.id_note, n.note_finale, n.commentaire AS commentaire_final
            FROM isms.soutenance s
            JOIN isms.memoire m ON m.id_memoire = s.id_memoire
            JOIN isms.etudiant e ON e.id_etudiant = m.id_etudiant
            JOIN isms.jury j ON j.id_jury = s.id_jury
            JOIN isms.salle sa ON sa.id_salle = s.id_salle
            LEFT JOIN isms.note n ON n.id_soutenance = s.id_soutenance
            WHERE s.id_soutenance = %s AND s.id_annee = %s
            LIMIT 1
        """, [soutenance_id, annee_id])
        base_rows = dictfetchall(cur)

    if not base_rows:
        messages.error(request, "Soutenance introuvable (ou pas dans l'année sélectionnée).")
        return redirect("gestion:note_list")

    base = base_rows[0]
    jury_id = base["id_jury"]
    memoire_id = base["id_memoire"]

    # Encadrements
    with connection.cursor() as cur:
        cur.execute("""
            SELECT r.id_responsable, r.nom, r.prenom, r.email, en.encadrement
            FROM isms.encadrement en
            JOIN isms.responsable r ON r.id_responsable = en.id_responsable
            WHERE en.id_memoire = %s
            ORDER BY en.encadrement, r.nom, r.prenom
        """, [memoire_id])
        encadrements = dictfetchall(cur)

    # Membres du jury + rôles (M2M responsable_role)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT
                r.id_responsable,
                r.nom, r.prenom, r.email,
                COALESCE(string_agg(DISTINCT ro.code, ', '), '')     AS role_code,
                COALESCE(string_agg(DISTINCT ro.libelle, ', '), '')  AS role_libelle
            FROM isms.composition_jury cj
            JOIN isms.responsable r ON r.id_responsable = cj.id_responsable
            LEFT JOIN isms.responsable_role rr ON rr.id_responsable = r.id_responsable
            LEFT JOIN isms.role ro ON ro.id_role = rr.id_role
            WHERE cj.id_jury = %s
            GROUP BY r.id_responsable, r.nom, r.prenom, r.email
            ORDER BY r.nom, r.prenom
        """, [jury_id])
        jury_membres = dictfetchall(cur)

    # Notes jury existantes
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id_responsable, note, commentaire
            FROM isms.note_jury
            WHERE id_soutenance = %s
        """, [soutenance_id])
        note_jury_rows = dictfetchall(cur)

    note_jury_map = {r["id_responsable"]: r for r in note_jury_rows}

    # ✅ IMPORTANT : construire jury_items pour le template
    jury_items = [{"m": m, "nj": note_jury_map.get(m["id_responsable"])} for m in jury_membres]

    # Moyenne harmonisée
    with connection.cursor() as cur:
        cur.execute("""
            SELECT ROUND(AVG(note)::numeric, 2)
            FROM isms.note_jury
            WHERE id_soutenance = %s
        """, [soutenance_id])
        moyenne = cur.fetchone()[0]

    return render(request, "gestion/notes/detail.html", {
        "base": base,
        "encadrements": encadrements,
        "jury_membres": jury_membres,
        "jury_items": jury_items,          # ✅ AJOUT
        "note_jury_map": note_jury_map,    # optionnel (plus utilisé par template actuel)
        "moyenne": moyenne,
    })



@superadmin_required
@admin_year_required
def note_jury_save(request, soutenance_id):
    if request.method != "POST":
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    annee_id = request.session.get("annee_id")

    id_responsable = request.POST.get("id_responsable")
    note = request.POST.get("note")
    commentaire = (request.POST.get("commentaire") or "").strip()

    if not id_responsable or note is None or note == "":
        messages.error(request, "Veuillez saisir la note du membre.")
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    try:
        # check membre appartient au jury de la soutenance
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 1
                FROM isms.soutenance s
                JOIN isms.composition_jury cj ON cj.id_jury = s.id_jury
                WHERE s.id_soutenance = %s
                  AND s.id_annee = %s
                  AND cj.id_responsable = %s
                LIMIT 1
            """, [soutenance_id, annee_id, id_responsable])
            if not cur.fetchone():
                messages.error(request, "Ce responsable n'est pas membre du jury de cette soutenance.")
                return redirect("gestion:note_detail", soutenance_id=soutenance_id)

        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO isms.note_jury (id_soutenance, id_responsable, note, commentaire)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (id_soutenance, id_responsable)
                DO UPDATE SET note = EXCLUDED.note,
                              commentaire = EXCLUDED.commentaire
            """, [soutenance_id, id_responsable, note, commentaire])

        messages.success(request, "Note du membre enregistrée.")
    except IntegrityError:
        messages.error(request, "Note invalide (doit être entre 0 et 20).")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:note_detail", soutenance_id=soutenance_id)


@superadmin_required
@admin_year_required
def note_jury_delete(request, soutenance_id, responsable_id):
    if request.method != "POST":
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    try:
        with connection.cursor() as cur:
            cur.execute("""
                DELETE FROM isms.note_jury
                WHERE id_soutenance=%s AND id_responsable=%s
            """, [soutenance_id, responsable_id])
        messages.success(request, "Note du membre supprimée.")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:note_detail", soutenance_id=soutenance_id)




    if request.method != "POST":
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    annee_id = request.session.get("annee_id")

    id_responsable = request.POST.get("id_responsable")
    note = request.POST.get("note")
    commentaire = (request.POST.get("commentaire") or "").strip()

    if not id_responsable or note is None or note == "":
        messages.error(request, "Veuillez remplir la note du membre.")
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    try:
        # Vérifier que ce responsable appartient au jury de cette soutenance
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 1
                FROM isms.soutenance s
                JOIN isms.composition_jury cj ON cj.id_jury = s.id_jury
                WHERE s.id_soutenance = %s
                  AND s.id_annee = %s
                  AND cj.id_responsable = %s
                LIMIT 1
            """, [soutenance_id, annee_id, id_responsable])
            ok = cur.fetchone()

        if not ok:
            messages.error(request, "Ce responsable n'est pas membre du jury de cette soutenance.")
            return redirect("gestion:note_detail", soutenance_id=soutenance_id)

        with connection.cursor() as cur:
            # Upsert (PK composite id_soutenance, id_responsable)
            cur.execute("""
                INSERT INTO isms.note_jury (id_soutenance, id_responsable, note, commentaire)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (id_soutenance, id_responsable)
                DO UPDATE SET note = EXCLUDED.note, commentaire = EXCLUDED.commentaire
            """, [soutenance_id, id_responsable, note, commentaire])

        messages.success(request, "Note du membre enregistrée.")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:note_detail", soutenance_id=soutenance_id)

@superadmin_required
@admin_year_required
def note_final_save(request, soutenance_id):
    if request.method != "POST":
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    annee_id = request.session.get("annee_id")

    note_finale = (request.POST.get("note_finale") or "").strip()
    commentaire = (request.POST.get("commentaire_final") or "").strip()

    try:
        # Vérifier soutenance dans l'année
        with connection.cursor() as cur:
            cur.execute("SELECT 1 FROM isms.soutenance WHERE id_soutenance=%s AND id_annee=%s", [soutenance_id, annee_id])
            if not cur.fetchone():
                messages.error(request, "Soutenance introuvable pour l'année sélectionnée.")
                return redirect("gestion:note_list")

        # si vide => moyenne
        if note_finale == "":
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT ROUND(AVG(note)::numeric, 2)
                    FROM isms.note_jury
                    WHERE id_soutenance=%s
                """, [soutenance_id])
                avg = cur.fetchone()[0]
            if avg is None:
                messages.error(request, "Impossible de valider : aucune note jury saisie.")
                return redirect("gestion:note_detail", soutenance_id=soutenance_id)
            note_finale = avg

        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO isms.note (note_finale, commentaire, id_soutenance)
                VALUES (%s, %s, %s)
                ON CONFLICT (id_soutenance)
                DO UPDATE SET note_finale = EXCLUDED.note_finale,
                              commentaire = EXCLUDED.commentaire
            """, [note_finale, commentaire, soutenance_id])

        messages.success(request, "Note finale enregistrée.")
    except IntegrityError:
        messages.error(request, "Note finale invalide (doit être entre 0 et 20).")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:note_detail", soutenance_id=soutenance_id)

@superadmin_required
@admin_year_required
def note_final_delete(request, soutenance_id):
    if request.method != "POST":
        return redirect("gestion:note_detail", soutenance_id=soutenance_id)

    annee_id = request.session.get("annee_id")

    try:
        with connection.cursor() as cur:
            # sécuriser année
            cur.execute("""
                DELETE FROM isms.note n
                USING isms.soutenance s
                WHERE n.id_soutenance = s.id_soutenance
                  AND n.id_soutenance = %s
                  AND s.id_annee = %s
            """, [soutenance_id, annee_id])
        messages.success(request, "Note finale supprimée.")
    except DatabaseError as e:
        messages.error(request, "Erreur base de données : " + str(e))

    return redirect("gestion:note_detail", soutenance_id=soutenance_id)
